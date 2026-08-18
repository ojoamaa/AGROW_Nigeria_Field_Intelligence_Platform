import os
import socket
import qrcode
from io import BytesIO
from datetime import datetime
import hmac
import hashlib
import time

import pandas as pd
import plotly.express as px
import streamlit as st
from streamlit_geolocation import streamlit_geolocation
from dotenv import load_dotenv
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image, ImageDraw, ImageFont
import re
from urllib.parse import quote
from sqlalchemy import text

from nigeria_lga_data import NIGERIA_LGA_MAP
from core.db import get_connection, init_db, get_engine, DB_BACKEND
from core.config import AGENT_PHOTO_DIR
from core.settings import invite_code_matches, get_invite_source, get_organisation_invite_code
from services.log_service import log_action
from services.farmer_service import (
    fetch_farmers,
    farmer_exists_today,
    insert_farmer,
    save_to_offline_queue,
)
from services.user_service import (
    fetch_user,
    insert_user,
    update_user_password,
    fetch_all_agents,
    generate_agent_id_db,
)
from services.market_service import (
    create_market_listing,
    fetch_market_listings,
    update_listing_status,
    create_input_product,
    fetch_input_products,
    update_input_status,
    create_market_enquiry,
    fetch_market_enquiries,
)

load_dotenv()


def ensure_field_sync_receipts_table():
    """Central audit trail written by AGROW Field after reconnect/sync."""
    try:
        engine = get_engine()
        id_column = "INTEGER PRIMARY KEY AUTOINCREMENT" if engine.dialect.name == "sqlite" else "BIGSERIAL PRIMARY KEY"
        ddl = f"""
        CREATE TABLE IF NOT EXISTS field_sync_receipts (
            id {id_column},
            farmer_id TEXT NOT NULL,
            agent_id TEXT NOT NULL,
            sync_status TEXT NOT NULL,
            received_at TEXT NOT NULL,
            error_message TEXT
        )
        """
        with engine.begin() as conn:
            conn.execute(text(ddl))
            # Collapse historical duplicate successful receipts and enforce
            # one accepted successful receipt per farmer + agent.
            conn.execute(text("""
                DELETE FROM field_sync_receipts
                WHERE UPPER(sync_status) = 'SYNCED'
                  AND id NOT IN (
                        SELECT MIN(id)
                        FROM field_sync_receipts
                        WHERE UPPER(sync_status) = 'SYNCED'
                        GROUP BY farmer_id, agent_id
                  )
            """))
            conn.execute(text("""
                CREATE UNIQUE INDEX IF NOT EXISTS ux_field_sync_success
                ON field_sync_receipts (farmer_id, agent_id)
                WHERE UPPER(sync_status) = 'SYNCED'
            """))
    except Exception:
        pass


def fetch_field_sync_receipts(limit=200):
    ensure_field_sync_receipts_table()
    try:
        with get_engine().connect() as conn:
            rows = conn.execute(text("""
                SELECT farmer_id, agent_id, sync_status, received_at, error_message
                FROM field_sync_receipts
                ORDER BY id DESC
                LIMIT :limit
            """), {"limit": int(limit)}).mappings().all()
        return pd.DataFrame([dict(r) for r in rows])
    except Exception:
        return pd.DataFrame()

APP_BASE_URL = os.getenv(
    "AGROW_LOCAL_BASE_URL",
    os.getenv(
        "APP_BASE_URL",
        "https://agrow-nigeria-field-intelligence-platform.onrender.com",
    ),
).rstrip("/")


def get_effective_app_base_url() -> str:
    """Return a QR-reachable base URL for production or local phone testing."""
    configured = (APP_BASE_URL or "http://localhost:8501").rstrip("/")
    if "localhost" not in configured and "127.0.0.1" not in configured:
        return configured

    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8", 80))
        lan_ip = sock.getsockname()[0]
        sock.close()
        return configured.replace("localhost", lan_ip).replace("127.0.0.1", lan_ip)
    except Exception:
        return configured


def seed_default_users():
    admin_user = fetch_user("admin")
    if not admin_user:
        insert_user(
            "admin",
            "agrow2026",
            "admin",
            "System Administrator",
            "08000000000",
            "00000000000",
            "FCT",
            "Abuja Municipal",
            "admin@datadev.local"
        )


def seed_demo_agents():
    if not fetch_user("FCT-01"):
        insert_user(
            "FCT-01",
            "agent2026",
            "agent",
            "Demo FCT Agent",
            "08011111111",
            "11111111111",
            "FCT",
            "Abaji",
            "fctagent@datadev.local"
        )

    if not fetch_user("KN-01"):
        insert_user(
            "KN-01",
            "agent2026",
            "agent",
            "Demo Kano Agent",
            "08022222222",
            "22222222222",
            "Kano",
            "Nassarawa",
            "kanoagent@datadev.local"
        )


def seed_demo_data():
    df = fetch_farmers()

    if not df.empty:
        return

    demo_rows = [
        {
            "Farmer_ID": "AG-202604220001",
            "Registration_Date": "2026-04-22 08:15:00",
            "Agent_ID": "FCT-01",
            "Farmer_Full_Name": "Musa Ibrahim",
            "Gender": "Male",
            "Date_of_Birth": "1988-05-12",
            "Phone_Number": "08031234567",
            "Alternate_Phone": "08141234567",
            "Email_Address": "musa@example.com",
            "NIN": "12345678901",
            "State": "FCT",
            "LGA": "Abaji",
            "Ward": "Ward 1",
            "Community_Village": "Yaba",
            "Residential_Address": "Abaji, FCT",
            "Primary_Crop": "Rice",
            "Secondary_Crop": "Maize",
            "Farm_Size_Ha": 3.5,
            "Input_Distributed": "Rice Seeds, NPK, Urea",
            "Quantity_Units": 3,
            "NIN_Status": "Verified",
            "ID_Type": "NIN Slip",
            "ID_Number": "NIN-001",
            "Latitude": 8.4732,
            "Longitude": 6.9421,
            "Enumerator_Remarks": "Demo registered farmer",
            "Photo_Path": "",
            "Photo_Status": "No Photo",
        },
        {
            "Farmer_ID": "AG-202604220002",
            "Registration_Date": "2026-04-22 08:35:00",
            "Agent_ID": "FCT-01",
            "Farmer_Full_Name": "Aisha Bello",
            "Gender": "Female",
            "Date_of_Birth": "1992-09-18",
            "Phone_Number": "08039876543",
            "Alternate_Phone": "",
            "Email_Address": "aisha@example.com",
            "NIN": "12345678902",
            "State": "FCT",
            "LGA": "Gwagwalada",
            "Ward": "Ward 2",
            "Community_Village": "Paiko",
            "Residential_Address": "Gwagwalada, FCT",
            "Primary_Crop": "Maize",
            "Secondary_Crop": "Soybean",
            "Farm_Size_Ha": 2.0,
            "Input_Distributed": "Maize Seeds, Organic Fertilizer",
            "Quantity_Units": 2,
            "NIN_Status": "Pending",
            "ID_Type": "Voter Card",
            "ID_Number": "PVC-002",
            "Latitude": 8.9390,
            "Longitude": 7.0819,
            "Enumerator_Remarks": "Awaiting NIN verification",
            "Photo_Path": "",
            "Photo_Status": "No Photo",
        },
        {
            "Farmer_ID": "AG-202604220003",
            "Registration_Date": "2026-04-22 09:05:00",
            "Agent_ID": "KN-01",
            "Farmer_Full_Name": "Sani Yakubu",
            "Gender": "Male",
            "Date_of_Birth": "1985-03-03",
            "Phone_Number": "08052345678",
            "Alternate_Phone": "",
            "Email_Address": "sani@example.com",
            "NIN": "12345678903",
            "State": "Kano",
            "LGA": "Nassarawa",
            "Ward": "Ward 3",
            "Community_Village": "Tudun Wada",
            "Residential_Address": "Kano",
            "Primary_Crop": "Rice",
            "Secondary_Crop": "Groundnut",
            "Farm_Size_Ha": 4.2,
            "Input_Distributed": "Rice Seeds, NPK, Herbicide, Sprayer",
            "Quantity_Units": 4,
            "NIN_Status": "Verified",
            "ID_Type": "Driver's License",
            "ID_Number": "DL-003",
            "Latitude": 11.9790,
            "Longitude": 8.5214,
            "Enumerator_Remarks": "Demo record from Kano cluster",
            "Photo_Path": "",
            "Photo_Status": "No Photo",
        },
    ]

    for row in demo_rows:
        insert_farmer(row)


init_db()
seed_default_users()
seed_demo_agents()

# Non-secret startup diagnostics. This helps identify configuration mismatches
# without exposing invite codes, database credentials, or signing keys.
STARTUP_INVITE_SOURCE = get_invite_source()
seed_demo_data()

if "farmer_form_version" not in st.session_state:
    st.session_state["farmer_form_version"] = 0

form_v = st.session_state["farmer_form_version"]
# =========================================================
# 1. PAGE CONFIG
# =========================================================
st.set_page_config(
    page_title="AGROW — Agricultural Geographic Registration & Operations Workspace",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =========================================================
# 2. ENV / DATABASE
# =========================================================


# =========================================================
# 3. BRANDING / COLORS
# =========================================================
PRIMARY_GREEN = "#1B5E20"
OFFICIAL_BLUE = "#004B87"
LIGHT_BG = "#F8F9FA"
TEXT_GREY = "#5F6368"

LOGO_CANDIDATES = [
    "proposal_logo.png",
    "logo.png",
    "static/proposal_logo.png",
    "static/logo.png",
]

# =========================================================
# 4. GLOBAL STYLING
# =========================================================
st.markdown(
    f"""
    <style>
    .stApp {{
        background-color: {LIGHT_BG};
    }}

    [data-testid="stSidebar"] {{
        background-color: #FFFFFF;
        border-right: 3px solid {PRIMARY_GREEN};
        overflow-y: auto !important;
    }}

    section[data-testid="stSidebar"] {{
        width: 320px !important;
    }}

    @media (max-width: 768px) {{
        section[data-testid="stSidebar"] {{
            width: 68vw !important;
            min-width: 68vw !important;
            max-width: 68vw !important;
        }}

        .block-container {{
            padding-top: 1rem !important;
            padding-left: 0.8rem !important;
            padding-right: 0.8rem !important;
            padding-bottom: 1rem !important;
        }}

        .main-title {{
            font-size: 16px !important;
            line-height: 1.25 !important;
            margin-bottom: 2px !important;
        }}

        .sub-title {{
            font-size: 10px !important;
            line-height: 1.3 !important;
            margin-bottom: 4px !important;
        }}

        .small-note {{
            font-size: 9px !important;
            line-height: 1.3 !important;
            margin-bottom: 8px !important;
        }}

        .top-title {{
            font-size: 11px !important;
            line-height: 1.2 !important;
        }}

        [data-testid="stMetric"] {{
            padding: 10px !important;
            border-radius: 10px !important;
        }}

        [data-testid="stMetricValue"] {{
            font-size: 20px !important;
        }}

        .footer {{
            display: none !important;
        }}

        .stButton > button,
        .stDownloadButton > button {{
            width: 100% !important;
            min-height: 44px !important;
            font-size: 14px !important;
        }}

        .stTextInput input,
        .stNumberInput input,
        .stDateInput input,
        .stTextArea textarea {{
            font-size: 14px !important;
        }}

        .auth-card {{
            max-width: 100% !important;
            padding: 12px !important;
            border-radius: 10px !important;
            margin-top: 6px !important;
        }}
    }}

    [data-testid="stSidebar"]::-webkit-scrollbar {{
        width: 10px;
    }}

    [data-testid="stSidebar"]::-webkit-scrollbar-track {{
        background: #F1F3F5;
        border-radius: 10px;
    }}

    [data-testid="stSidebar"]::-webkit-scrollbar-thumb {{
        background: #B0B8C1;
        border-radius: 10px;
    }}

    [data-testid="stSidebar"]::-webkit-scrollbar-thumb:hover {{
        background: #8C959F;
    }}

    [data-testid="stMetric"] {{
        background: white !important;
        border-left: 8px solid {PRIMARY_GREEN} !important;
        border-radius: 14px !important;
        padding: 16px !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.06) !important;
    }}

    [data-testid="stMetricValue"] {{
        color: {OFFICIAL_BLUE} !important;
        font-size: 28px !important;
        font-weight: 800 !important;
    }}

    .auth-card {{
        padding: 18px 24px 24px 24px;
        background: white;
        border-radius: 16px;
        border-top: 8px solid {PRIMARY_GREEN};
        box-shadow: 0 12px 30px rgba(0,0,0,0.10);
        max-width: 760px;
        margin: auto;
    }}

    .top-title {{
        text-align: center;
        font-size: 14px;
        font-weight: 800;
        color: {PRIMARY_GREEN};
        letter-spacing: 0.3px;
    }}

    .main-title {{
        text-align: center;
        font-size: 34px;
        font-weight: 900;
        color: {OFFICIAL_BLUE};
        margin-top: 2px;
        margin-bottom: 4px;
        line-height: 1.2;
    }}

    .sub-title {{
        text-align: center;
        font-size: 15px;
        font-weight: 700;
        color: {PRIMARY_GREEN};
        margin-bottom: 6px;
        line-height: 1.4;
    }}

    .small-note {{
        text-align: center;
        font-size: 13px;
        color: {TEXT_GREY};
        margin-bottom: 14px;
        line-height: 1.4;
    }}

    label,
    .stSelectbox label,
    .stMultiSelect label,
    .stTextInput label,
    .stDateInput label,
    .stNumberInput label,
    .stTextArea label {{
        font-weight: 700 !important;
        color: #1f1f1f !important;
    }}

    .stTextInput input,
    .stNumberInput input,
    .stDateInput input,
    .stTextArea textarea {{
        border: 1.5px solid #D0D7DE !important;
        border-radius: 10px !important;
        background-color: #FFFFFF !important;
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        caret-color: #111111 !important;
        padding: 0.65rem !important;
        font-size: 16px !important;
        opacity: 1 !important;
    }}

    div[data-baseweb="input"] input {{
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        caret-color: #111111 !important;
        background-color: #FFFFFF !important;
    }}

    textarea {{
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        caret-color: #111111 !important;
        background-color: #FFFFFF !important;
    }}

    input::placeholder,
    textarea::placeholder {{
        color: #777777 !important;
        opacity: 1 !important;
    }}

    div[data-baseweb="select"] > div {{
        border: 1.5px solid #D0D7DE !important;
        border-radius: 10px !important;
        background-color: #FAFBFC !important;
    }}

    .sidebar-form-note {{
        background: #F4F8F4;
        border-left: 4px solid {PRIMARY_GREEN};
        padding: 10px 12px;
        border-radius: 8px;
        font-size: 13px;
        margin-bottom: 10px;
        color: #2f3b2f;
    }}

    .stButton > button,
    .stDownloadButton > button {{
        border-radius: 10px !important;
        font-weight: 700 !important;
        min-height: 44px !important;
    }}

    .stApp::after {{
        content: "DataDev Limited";
        position: fixed;
        bottom: 90px;
        right: 20px;
        font-size: 36px;
        color: rgba(0, 0, 0, 0.03);
        transform: rotate(-20deg);
        z-index: 0;
        pointer-events: none;
    }}

    /* =====================================================
       PRODUCTION UI POLISH — visibility, touch targets, mobile
       ===================================================== */
    .stButton > button,
    .stDownloadButton > button,
    .stLinkButton > a {{
        min-height: 48px !important;
        border-radius: 10px !important;
        font-weight: 800 !important;
        font-size: 15px !important;
        border: 1.5px solid {OFFICIAL_BLUE} !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.08) !important;
    }}

    .stButton > button[kind="primary"],
    .stFormSubmitButton > button[kind="primary"] {{
        background: {OFFICIAL_BLUE} !important;
        color: #FFFFFF !important;
        border-color: {OFFICIAL_BLUE} !important;
    }}

    .stDownloadButton > button {{
        background: {PRIMARY_GREEN} !important;
        color: #FFFFFF !important;
        border-color: {PRIMARY_GREEN} !important;
    }}

    .stLinkButton > a {{
        background: #FFFFFF !important;
        color: {OFFICIAL_BLUE} !important;
    }}

    /* Feature navigation: make every workspace module visually obvious. */
    button[data-baseweb="tab"] {{
        min-height: 52px !important;
        padding: 9px 16px !important;
        font-weight: 850 !important;
        font-size: 14px !important;
        color: {OFFICIAL_BLUE} !important;
        background: #F4F8FB !important;
        border: 1.5px solid #B8CBD9 !important;
        border-radius: 11px !important;
        box-shadow: 0 2px 5px rgba(0,75,135,0.07) !important;
        margin-bottom: 4px !important;
    }}

    button[data-baseweb="tab"] p {{
        color: inherit !important;
        font-weight: 850 !important;
    }}

    button[data-baseweb="tab"]:hover {{
        background: #E7F1F8 !important;
        border-color: {OFFICIAL_BLUE} !important;
    }}

    button[data-baseweb="tab"][aria-selected="true"] {{
        color: #FFFFFF !important;
        background: {OFFICIAL_BLUE} !important;
        border-color: {OFFICIAL_BLUE} !important;
        box-shadow: 0 4px 10px rgba(0,75,135,0.20) !important;
    }}

    button[data-baseweb="tab"][aria-selected="true"] p {{
        color: #FFFFFF !important;
    }}

    div[data-baseweb="tab-list"] {{
        gap: 7px !important;
        overflow-x: auto !important;
        scrollbar-width: thin;
        padding: 3px 2px 7px 2px !important;
    }}

    [data-testid="stExpander"] {{
        background: #FFFFFF !important;
        border: 1px solid #DDE3E8 !important;
        border-radius: 12px !important;
    }}

    [data-testid="stVerticalBlockBorderWrapper"] {{
        background: #FFFFFF;
        border-radius: 12px;
    }}

    @media (max-width: 768px) {{
        .stButton > button,
        .stDownloadButton > button,
        .stLinkButton > a {{
            min-height: 50px !important;
            font-size: 14px !important;
        }}

        button[data-baseweb="tab"] {{
            min-width: max-content !important;
            min-height: 46px !important;
            padding: 7px 11px !important;
            font-size: 13px !important;
        }}

        [data-testid="column"] {{
            min-width: 0 !important;
        }}

        h1 {{ font-size: 1.55rem !important; }}
        h2 {{ font-size: 1.35rem !important; }}
        h3 {{ font-size: 1.15rem !important; }}
    }}

    .footer {{
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        background-color: white;
        text-align: center;
        padding: 10px;
        font-weight: 700;
        border-top: 1px solid #ddd;
        z-index: 999;
        font-size: 12px;
    }}

    /* Production visibility refinements */
    div.stButton > button, div.stDownloadButton > button, a[data-testid="stBaseLinkButton-primary"] {{
        min-height: 44px !important;
        font-weight: 800 !important;
        border-radius: 10px !important;
    }}
    button[kind="primary"], a[data-testid="stBaseLinkButton-primary"] {{
        background: #0069B4 !important;
        color: #FFFFFF !important;
        border: 1px solid #00558F !important;
    }}
    [data-baseweb="tab-list"] {{
        gap: 6px !important;
        flex-wrap: wrap !important;
    }}
    [data-baseweb="tab"] {{
        background: #EEF5FA !important;
        border: 1px solid #C7D9E8 !important;
        border-radius: 8px 8px 0 0 !important;
        padding: 9px 12px !important;
        font-weight: 750 !important;
        color: #163A54 !important;
    }}
    [aria-selected="true"][data-baseweb="tab"] {{
        background: #E6F4EA !important;
        color: #145A32 !important;
        border-bottom: 3px solid #1B5E20 !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

def generate_qr_code(url):
    qr = qrcode.make(url)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer

def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "").strip())
    return cleaned.strip("_") or "farmer"

def build_farmer_qr_identity_image(farmer, verification_url: str):
    width, height = 900, 1100
    card = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(card)
    font = ImageFont.load_default()
    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(verification_url); qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_img.thumbnail((620, 620))
    draw.rectangle((20, 20, width-20, height-20), outline=(0,83,141), width=8)
    details = [
        "AGROW FARMER IDENTIFICATION QR",
        f"Name: {farmer.get('farmer_full_name', '-')}",
        f"Farmer ID: {farmer.get('farmer_id', '-')}",
        f"State: {normalize_state_name(farmer.get('state', '-'))}",
        f"LGA: {farmer.get('lga', '-')}",
        f"Community: {farmer.get('community', '-')}",
        f"Primary Crop: {farmer.get('primary_crop', '-')}",
    ]
    y=55
    for line in details:
        draw.text((55,y), line, fill=(0,83,141) if y==55 else "black", font=font); y += 40
    card.paste(qr_img, ((width-qr_img.width)//2, 360))
    draw.text((55,1020), "Scan to verify this farmer's AGROW record.", fill=(55,55,55), font=font)
    out=BytesIO(); card.save(out, format="PNG"); out.seek(0); return out

STATE_GPS_ENVELOPES = {
    "Jigawa": (10.5, 13.2, 8.0, 10.7),
    "Kano": (10.4, 12.7, 7.5, 9.7),
    "Kaduna": (9.0, 11.8, 6.0, 9.0),
    "FCT": (8.2, 9.4, 6.6, 7.8),
    "Federal Capital Territory": (8.2, 9.4, 6.6, 7.8),
}

def gps_matches_assigned_state(state_name, latitude, longitude):
    bounds = STATE_GPS_ENVELOPES.get(str(state_name or "").strip())
    if not bounds: return None
    a,b,c,d = bounds
    return a <= latitude <= b and c <= longitude <= d

QR_SECRET_KEY = os.getenv("QR_SECRET_KEY", "change-this-secret-key")

def generate_qr_signature(farmer_id):
    message = farmer_id.encode()
    secret = QR_SECRET_KEY.encode()
    return hmac.new(secret, message, hashlib.sha256).hexdigest()


def verify_qr_signature(farmer_id, sig):
    expected_sig = generate_qr_signature(farmer_id)
    return hmac.compare_digest(expected_sig, sig)

# =========================================================
# 5. HELPERS
# =========================================================
def get_logo_path():
    for path in LOGO_CANDIDATES:
        if os.path.exists(path):
            return path
    return None


def normalize_state_name(state_name: str) -> str:
    if state_name == "FCT":
        return "Federal Capital Territory"
    return state_name

FARMER_PHOTO_DIR = os.path.join("uploads", "farmers")


def save_uploaded_photo(photo_file, folder: str, filename: str) -> str:
    """Save an uploaded/camera image and return a portable project-relative path.

    AGROW stores the relative path in the farmer record rather than a Windows
    absolute path. This keeps the same database record usable locally and on
    Render. The write is verified before the farmer record is committed.
    """
    os.makedirs(folder, exist_ok=True)
    file_path = os.path.normpath(os.path.join(folder, filename))

    photo_bytes = photo_file.getvalue() if hasattr(photo_file, "getvalue") else photo_file.getbuffer()
    with open(file_path, "wb") as f:
        f.write(photo_bytes)

    if not os.path.isfile(file_path) or os.path.getsize(file_path) == 0:
        raise IOError(f"Farmer photo could not be saved correctly: {file_path}")

    return file_path.replace("\\", "/")


def resolve_farmer_photo_path(photo_path="", farmer_id="") -> str:
    """Resolve farmer photos across local Windows and Render/Linux deployments.

    Older records may contain a path created on another machine. We therefore
    try the stored path first, then the current uploads/farmers directory using
    both the stored basename and the Farmer ID. No database mutation is needed.
    """
    raw_path = str(photo_path or "").strip()
    farmer_id = str(farmer_id or "").strip()

    candidates = []
    if raw_path and raw_path.lower() != "nan":
        candidates.append(raw_path)
        candidates.append(os.path.normpath(raw_path))
        basename = os.path.basename(raw_path.replace("\\", "/"))
        if basename:
            candidates.append(os.path.join(FARMER_PHOTO_DIR, basename))

    if farmer_id:
        for ext in (".jpg", ".jpeg", ".png", ".webp"):
            candidates.append(os.path.join(FARMER_PHOTO_DIR, f"{farmer_id}{ext}"))

    seen = set()
    for candidate in candidates:
        normalized = os.path.normpath(str(candidate))
        if normalized in seen:
            continue
        seen.add(normalized)
        if os.path.isfile(normalized) and os.path.getsize(normalized) > 0:
            return normalized

    return ""

def get_state_prefix(state_name: str) -> str:
    special_map = {
        "Abia": "AB",
        "Adamawa": "AD",
        "Akwa Ibom": "AI",
        "Anambra": "AN",
        "Bauchi": "BC",
        "Bayelsa": "BY",
        "Benue": "BN",
        "Borno": "BO",
        "Cross River": "CR",
        "Delta": "DT",
        "Ebonyi": "EB",
        "Edo": "ED",
        "Ekiti": "EK",
        "Enugu": "EN",
        "FCT": "FC",
        "Federal Capital Territory": "FC",
        "Gombe": "GM",
        "Imo": "IM",
        "Jigawa": "JG",
        "Kaduna": "KD",
        "Kano": "KN",
        "Katsina": "KT",
        "Kebbi": "KB",
        "Kogi": "KG",
        "Kwara": "KW",
        "Lagos": "LG",
        "Nasarawa": "NS",
        "Niger": "NG",
        "Ogun": "OG",
        "Ondo": "OD",
        "Osun": "OS",
        "Oyo": "OY",
        "Plateau": "PL",
        "Rivers": "RV",
        "Sokoto": "SK",
        "Taraba": "TR",
        "Yobe": "YB",
        "Zamfara": "ZF",
    }
    return special_map.get(state_name, state_name[:2].upper())


def valid_phone(phone):
    return phone.isdigit() and 10 <= len(phone) <= 11


def valid_nin(nin: str) -> bool:
    return nin.isdigit() and len(nin) == 11

def valid_id_number(id_number: str) -> bool:
    cleaned = id_number.strip()
    return 1 <= len(cleaned) <= 20

def valid_email(email: str) -> bool:
    return "@" in email and "." in email



def fetch_complete_farmer_registry() -> pd.DataFrame:
    """Read the authoritative farmers table directly for complete analysis export.

    This deliberately does not depend on the compact fetch_farmers() projection,
    because the export must contain every non-image registration field stored in
    the farmers table.
    """
    try:
        with get_engine().connect() as conn:
            result = conn.execute(text("SELECT * FROM farmers"))
            rows = result.mappings().all()
        return pd.DataFrame([dict(r) for r in rows])
    except Exception:
        # Safe fallback for environments where the full table cannot be queried.
        try:
            return fetch_farmers().copy()
        except Exception:
            return pd.DataFrame()


def _canonical_farmer_column_name(column_name: str) -> str:
    """Map database naming variants to the AGROW export schema."""
    raw = str(column_name or "").strip()
    key = re.sub(r"[^a-z0-9]+", "_", raw.lower()).strip("_")
    aliases = {
        "id": "Database_ID",
        "farmer_id": "Farmer_ID",
        "registration_date": "Registration_Date",
        "agent_id": "Agent_ID",
        "farmer_full_name": "Farmer_Full_Name",
        "gender": "Gender",
        "date_of_birth": "Date_of_Birth",
        "phone_number": "Phone_Number",
        "alternate_phone": "Alternate_Phone",
        "email_address": "Email_Address",
        "nin": "NIN",
        "state": "State",
        "lga": "LGA",
        "ward": "Ward",
        "community_village": "Community_Village",
        "community": "Community_Village",
        "village": "Community_Village",
        "residential_address": "Residential_Address",
        "address": "Residential_Address",
        "primary_crop": "Primary_Crop",
        "secondary_crop": "Secondary_Crop",
        "farm_size_ha": "Farm_Size_Ha",
        "farm_size": "Farm_Size_Ha",
        "input_distributed": "Input_Distributed",
        "inputs_distributed": "Input_Distributed",
        "support_delivered": "Input_Distributed",
        "input_support_delivered": "Input_Distributed",
        "quantity_units": "Quantity_Units",
        "input_quantity": "Quantity_Units",
        "quantity": "Quantity_Units",
        "nin_status": "NIN_Status",
        "id_type": "ID_Type",
        "id_number": "ID_Number",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "gps_accuracy": "GPS_Accuracy_M",
        "enumerator_remarks": "Enumerator_Remarks",
        "remarks": "Enumerator_Remarks",
        "photo_status": "Photo_Status",
        "photo_path": "Photo_Path",
        "photo_data": "Photo_Data",
        "photo_mime": "Photo_Mime",
    }
    return aliases.get(key, raw)


def _normalise_nigerian_phone(value) -> str:
    """Return a phone identifier as text and restore common Nigerian leading zeroes."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text_value = str(value).strip()
    if not text_value or text_value.lower() in {"nan", "none"}:
        return ""

    # Undo Excel/database numeric formatting where possible (e.g. 8.123E+09).
    if re.fullmatch(r"\d+(?:\.0+)?", text_value):
        text_value = text_value.split(".", 1)[0]
    elif re.fullmatch(r"\d+(?:\.\d+)?[eE][+-]?\d+", text_value):
        try:
            text_value = format(float(text_value), ".0f")
        except Exception:
            pass

    digits = re.sub(r"\D", "", text_value)
    if digits.startswith("234") and len(digits) == 13:
        digits = "0" + digits[3:]
    elif len(digits) == 10 and digits[0] in "789":
        digits = "0" + digits

    # Keep standard Nigerian mobile length, otherwise preserve cleaned content.
    if len(digits) in {10, 11, 13}:
        return digits
    return text_value

def clean_registry_export(df: pd.DataFrame) -> pd.DataFrame:
    """Export every farmer detail except raw/technical image content.

    Keeps all operational registration fields (including inputs distributed,
    quantities, location, identity, remarks and any future non-image fields),
    while excluding photo paths, MIME/base64/binary payloads and payload-only rows.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    # Canonicalise all naming variants first so complete database fields are retained.
    rename_all = {c: _canonical_farmer_column_name(c) for c in out.columns}
    out = out.rename(columns=rename_all)

    farmer_aliases = {
        "farmer_id": "Farmer_ID", "registration_date": "Registration_Date",
        "agent_id": "Agent_ID", "farmer_full_name": "Farmer_Full_Name",
        "gender": "Gender", "date_of_birth": "Date_of_Birth",
        "phone_number": "Phone_Number", "alternate_phone": "Alternate_Phone",
        "email_address": "Email_Address", "nin": "NIN", "state": "State",
        "lga": "LGA", "ward": "Ward", "community_village": "Community_Village",
        "residential_address": "Residential_Address", "primary_crop": "Primary_Crop",
        "secondary_crop": "Secondary_Crop", "farm_size_ha": "Farm_Size_Ha",
        "input_distributed": "Input_Distributed", "quantity_units": "Quantity_Units",
        "nin_status": "NIN_Status", "id_type": "ID_Type", "id_number": "ID_Number",
        "latitude": "Latitude", "longitude": "Longitude",
        "enumerator_remarks": "Enumerator_Remarks", "photo_status": "Photo_Status",
        "photo_path": "Photo_Path",
    }
    rename_map = {k: v for k, v in farmer_aliases.items() if k in out.columns and v not in out.columns}
    if rename_map:
        out = out.rename(columns=rename_map)

    # If legacy/new schema variants collapse to the same canonical name, coalesce them.
    if out.columns.duplicated().any():
        rebuilt = pd.DataFrame(index=out.index)
        for name in dict.fromkeys(out.columns):
            same = out.loc[:, out.columns == name]
            if same.shape[1] == 1:
                rebuilt[name] = same.iloc[:, 0]
            else:
                merged = same.iloc[:, 0]
                for i in range(1, same.shape[1]):
                    candidate = same.iloc[:, i]
                    empty = merged.isna() | merged.astype(str).str.strip().isin({"", "nan", "None"})
                    merged = merged.where(~empty, candidate)
                rebuilt[name] = merged
        out = rebuilt

    is_farmer_register = "Farmer_ID" in out.columns

    # Retain a harmless Yes/No photo indicator, but never export the image/path/payload.
    if is_farmer_register:
        status = out["Photo_Status"].fillna("").astype(str) if "Photo_Status" in out.columns else pd.Series("", index=out.index)
        path = out["Photo_Path"].fillna("").astype(str) if "Photo_Path" in out.columns else pd.Series("", index=out.index)
        captured = status.str.strip().str.lower().isin({"captured", "yes", "true", "1", "photo captured"}) | path.str.strip().replace({"nan": "", "None": ""}).ne("")
        out["Photo_Captured"] = captured.map({True: "Yes", False: "No"})

    # Drop ONLY image/binary technical columns. All other registration details remain.
    blocked_tokens = (
        "photo_path", "photo_data", "image_data", "photo_mime", "image_mime",
        "base64", "binary_data", "blob", "raw_image", "image_bytes", "photo_bytes"
    )
    drop_cols = [c for c in out.columns if any(token in str(c).lower() for token in blocked_tokens)]
    out = out.drop(columns=drop_cols, errors="ignore")

    if is_farmer_register:
        farmer_ids = out["Farmer_ID"].fillna("").astype(str).str.strip()
        out = out.loc[farmer_ids.ne("") & ~farmer_ids.str.lower().isin({"nan", "none"})].copy()
        if "Registration_Date" in out.columns:
            out = out.assign(_sort_dt=pd.to_datetime(out["Registration_Date"], errors="coerce")).sort_values("_sort_dt", ascending=False, na_position="last")
        out = out.drop_duplicates(subset=["Farmer_ID"], keep="first").drop(columns=["_sort_dt"], errors="ignore")

    def safe_text(value):
        if value is None:
            return ""
        try:
            if pd.isna(value): return ""
        except Exception:
            pass
        if isinstance(value, (bytes, bytearray, memoryview)): return ""
        value = str(value)
        low = value.lower().strip()
        if any(m in low for m in ("data:image/", ";base64,", "image/jpeg", "image/png", "image/webp", "application/octet-stream")):
            return ""
        value = value.replace("\r", " " ).replace("\n", " " ).replace("\t", " " )
        value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", value)
        return re.sub(r"\s{2,}", " ", value).strip()

    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].map(safe_text)

    text_columns = {"Farmer_ID", "Agent_ID", "Phone_Number", "Alternate_Phone", "NIN", "ID_Number", "Ward"}
    for col in text_columns.intersection(out.columns):
        out[col] = out[col].fillna("").astype(str).map(safe_text)

    for col in {"Phone_Number", "Alternate_Phone"}.intersection(out.columns):
        out[col] = out[col].map(_normalise_nigerian_phone)

    if "Registration_Date" in out.columns:
        parsed = pd.to_datetime(out["Registration_Date"], errors="coerce")
        out["Registration_Date"] = [dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(dt) else safe_text(raw) for dt, raw in zip(parsed, out["Registration_Date"])]
    if "Date_of_Birth" in out.columns:
        parsed = pd.to_datetime(out["Date_of_Birth"], errors="coerce")
        out["Date_of_Birth"] = [dt.strftime("%Y-%m-%d") if pd.notna(dt) else safe_text(raw) for dt, raw in zip(parsed, out["Date_of_Birth"])]

    return out.reset_index(drop=True)


def _excel_safe_csv_text(value):
    """Make numeric identifiers display as text when CSV is opened in Excel."""
    text_value = str(value or "").strip()
    if not text_value:
        return ""
    # Only wrap digit-only values. This avoids formula injection from arbitrary text.
    if text_value.isdigit():
        return f'="{text_value}"'
    return text_value


def table_to_csv_download(df: pd.DataFrame, filename: str, key: str | None = None):
    export_df = clean_registry_export(df)

    # CSV is a data-interchange format and Excel otherwise turns phone/NIN values
    # into numbers/scientific notation.  Wrap digit-only identity fields in a
    # safe Excel text formula so leading zeroes remain visible when double-clicked.
    csv_df = export_df.copy()
    excel_text_columns = {
        "Phone_Number", "Alternate_Phone", "NIN", "ID_Number"
    }
    for col in excel_text_columns.intersection(csv_df.columns):
        csv_df[col] = csv_df[col].map(_excel_safe_csv_text)

    csv_data = csv_df.to_csv(index=False, lineterminator="\r\n").encode("utf-8-sig")
    st.download_button(
        label=f"⬇️ Download CSV — {filename}",
        data=csv_data,
        file_name=filename,
        mime="text/csv",
        width="stretch",
        key=key or f"csv_download_{safe_filename(filename)}",
    )


def table_to_excel_download(df: pd.DataFrame, filename: str, key: str | None = None):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    export_df = clean_registry_export(df)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Farmer Register")
        ws = writer.book["Farmer Register"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F6B45")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        text_columns = {
            "Farmer_ID", "Agent_ID", "Phone_Number", "Alternate_Phone",
            "NIN", "ID_Number", "Ward",
        }

        for idx, col_name in enumerate(export_df.columns, start=1):
            letter = get_column_letter(idx)
            values = [str(col_name)] + [
                str(v) if v is not None else ""
                for v in export_df[col_name].head(200)
            ]
            max_len = min(max((len(v) for v in values), default=10) + 2, 42)
            ws.column_dimensions[letter].width = max(12, max_len)

            if col_name in text_columns:
                for cell in ws[letter][1:]:
                    # Force the actual stored value to a string, not merely the
                    # display format. This prevents 081... -> 8.1E+09.
                    cell.value = "" if cell.value is None else str(cell.value)
                    cell.number_format = "@"

            if col_name in {"Registration_Date", "Date_of_Birth"}:
                ws.column_dimensions[letter].width = max(
                    ws.column_dimensions[letter].width, 20
                )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # Friendly worksheet settings.
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 28

    output.seek(0)

    st.download_button(
        label=f"⬇️ Download Excel — {filename}",
        data=output.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        key=key or f"excel_download_{safe_filename(filename)}",
    )


def is_db_available() -> bool:
    try:
        from sqlalchemy import text
        with get_engine().connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False
    
def _fetch_scoped_dashboard_farmers(selected_state, role, user_id) -> pd.DataFrame:
    """Fetch a fresh Farmer Register view for live dashboard fragments."""
    fresh_df = fetch_farmers()
    if fresh_df is None or fresh_df.empty:
        return pd.DataFrame()

    if selected_state != "All Nigeria" and "state" in fresh_df.columns:
        fresh_df = fresh_df[fresh_df["state"] == selected_state]

    if role == "agent" and "agent_id" in fresh_df.columns:
        fresh_df = fresh_df[
            fresh_df["agent_id"].fillna("").astype(str).str.upper()
            == str(user_id or "").upper()
        ]

    return fresh_df.copy()


@st.fragment(run_every="10s")
def render_live_dashboard_overview(selected_state, role, user_id):
    """Keep headline dashboard counts current without refreshing/logging out the browser."""
    live_df = _fetch_scoped_dashboard_farmers(selected_state, role, user_id)

    total_beneficiaries = len(live_df)
    total_verified = len(live_df[live_df["nin_status"] == "Verified"]) if not live_df.empty and "nin_status" in live_df.columns else 0
    total_land = float(pd.to_numeric(live_df.get("farm_size_ha", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not live_df.empty else 0.0
    verification_rate = round((total_verified / total_beneficiaries) * 100, 1) if total_beneficiaries > 0 else 0.0

    registrations_today = 0
    pending_nin = 0
    rejected_nin = 0
    captured_photos = 0

    if not live_df.empty:
        today_str = datetime.now().strftime("%Y-%m-%d")
        if "registration_date" in live_df.columns:
            registrations_today = int(
                live_df["registration_date"].astype(str).str.startswith(today_str).sum()
            )
        if "nin_status" in live_df.columns:
            status = live_df["nin_status"].fillna("").astype(str)
            pending_nin = int(status.eq("Pending").sum())
            rejected_nin = int(status.eq("Rejected").sum())
        if "photo_status" in live_df.columns:
            captured_photos = int(
                live_df["photo_status"].fillna("").astype(str).eq("Captured").sum()
            )

    st.markdown("### Agricultural Operations Overview")
    row1 = st.columns(4)
    row1[0].metric("Registered Farmers", total_beneficiaries)
    row1[1].metric("Verified NIN", total_verified)
    row1[2].metric("Land Coverage (Ha)", f"{total_land:.1f}")
    row1[3].metric("Verification Rate", f"{verification_rate}%")

    row2 = st.columns(4)
    row2[0].metric("Registrations Today", registrations_today)
    row2[1].metric("Pending NIN", pending_nin)
    row2[2].metric("Rejected NIN", rejected_nin)
    row2[3].metric("Captured Photos", captured_photos)


def _build_agent_performance(farmers_df: pd.DataFrame, receipts_df: pd.DataFrame) -> pd.DataFrame:
    """Build registration + device-sync performance from fresh authoritative data."""
    receipt_perf = pd.DataFrame()
    register_perf = pd.DataFrame()
    today_date = pd.Timestamp.now().date()

    if receipts_df is not None and not receipts_df.empty:
        rp = receipts_df.copy()
        rp["received_at_dt"] = pd.to_datetime(rp.get("received_at"), errors="coerce")
        rp["sync_status_norm"] = rp["sync_status"].astype(str).str.upper()

        receipt_perf = (
            rp.groupby("agent_id", dropna=False)
            .agg(
                total_sync_receipts=("sync_status_norm", lambda s: int((s == "SYNCED").sum())),
                failed_syncs=("sync_status_norm", lambda s: int((s == "FAILED").sum())),
                last_activity=("received_at_dt", "max"),
            )
            .reset_index()
        )

        unique_perf = (
            rp[rp["sync_status_norm"].eq("SYNCED")]
            .groupby("agent_id", dropna=False)["farmer_id"]
            .nunique()
            .rename("unique_farmers_synced")
            .reset_index()
        )
        today_perf = (
            rp[
                rp["sync_status_norm"].eq("SYNCED")
                & rp["received_at_dt"].dt.date.eq(today_date)
            ]
            .groupby("agent_id", dropna=False)["farmer_id"]
            .nunique()
            .rename("synced_today")
            .reset_index()
        )
        receipt_perf = receipt_perf.merge(unique_perf, on="agent_id", how="left")
        receipt_perf = receipt_perf.merge(today_perf, on="agent_id", how="left")

    if farmers_df is not None and not farmers_df.empty and "agent_id" in farmers_df.columns:
        reg = farmers_df.copy()
        reg["registration_date_dt"] = pd.to_datetime(reg.get("registration_date"), errors="coerce")
        register_perf = (
            reg.groupby("agent_id", dropna=False)
            .agg(
                registered_farmers=("farmer_id", "nunique"),
                last_registration=("registration_date_dt", "max"),
            )
            .reset_index()
        )
        registered_today = (
            reg[reg["registration_date_dt"].dt.date.eq(today_date)]
            .groupby("agent_id", dropna=False)["farmer_id"]
            .nunique()
            .rename("registered_today")
            .reset_index()
        )
        register_perf = register_perf.merge(registered_today, on="agent_id", how="left")

    if not receipt_perf.empty and not register_perf.empty:
        agent_perf = receipt_perf.merge(register_perf, on="agent_id", how="outer")
    elif not receipt_perf.empty:
        agent_perf = receipt_perf.copy()
    else:
        agent_perf = register_perf.copy()

    if agent_perf.empty:
        return agent_perf

    numeric_cols = [
        "total_sync_receipts", "unique_farmers_synced", "synced_today",
        "failed_syncs", "registered_farmers", "registered_today",
    ]
    for col in numeric_cols:
        if col not in agent_perf.columns:
            agent_perf[col] = 0
        agent_perf[col] = agent_perf[col].fillna(0).astype(int)

    for col in ["last_activity", "last_registration"]:
        if col not in agent_perf.columns:
            agent_perf[col] = pd.NaT

    return agent_perf[
        [
            "agent_id", "registered_farmers", "registered_today",
            "unique_farmers_synced", "synced_today", "total_sync_receipts",
            "failed_syncs", "last_registration", "last_activity",
        ]
    ].sort_values(
        by=["registered_today", "registered_farmers", "unique_farmers_synced"],
        ascending=False,
    )


@st.fragment(run_every="10s")
def render_live_sync_monitor(selected_state, role, user_id):
    """Live central sync monitor and field-user summary, refreshed every 10 seconds."""
    st.markdown("### Central Sync Monitor")
    action_col, note_col = st.columns([1, 3])
    with action_col:
        if st.button("↻ Refresh Sync Monitor", key="refresh_central_sync_monitor", width="stretch"):
            st.rerun()
    with note_col:
        st.caption("Live field activity refreshes automatically every 10 seconds while this dashboard is open.")

    st.caption(
        "This monitor is an operational audit trail of AGROW Field device transmissions. "
        "Successful sync receipts are idempotent: one accepted receipt per farmer and agent. "
        "The Farmer Register remains the authoritative source for unique farmer registrations and registration dates."
    )

    receipts_df = fetch_field_sync_receipts()
    fresh_df = _fetch_scoped_dashboard_farmers(selected_state, role, user_id)

    total_sync_receipts = 0
    unique_farmers_synced = 0
    synced_today = 0
    field_failed_count = 0
    last_received = "—"

    if not receipts_df.empty:
        receipts_work = receipts_df.copy()
        receipts_work["received_at_dt"] = pd.to_datetime(receipts_work.get("received_at"), errors="coerce")
        synced_mask = receipts_work["sync_status"].astype(str).str.upper().eq("SYNCED")
        failed_mask = receipts_work["sync_status"].astype(str).str.upper().eq("FAILED")
        total_sync_receipts = int(synced_mask.sum())
        field_failed_count = int(failed_mask.sum())

        if "farmer_id" in receipts_work.columns:
            unique_farmers_synced = int(
                receipts_work.loc[synced_mask, "farmer_id"]
                .dropna().astype(str).replace("", pd.NA).dropna().nunique()
            )

        today_date = pd.Timestamp.now().date()
        today_success = receipts_work.loc[
            synced_mask & receipts_work["received_at_dt"].dt.date.eq(today_date)
        ]
        synced_today = int(
            today_success["farmer_id"].dropna().astype(str).nunique()
            if "farmer_id" in today_success.columns else len(today_success)
        )
        if "received_at" in receipts_work.columns:
            last_received = str(receipts_work.iloc[0]["received_at"] or "—")

    q1, q2, q3, q4 = st.columns(4)
    q1.metric("Total Sync Receipts", total_sync_receipts)
    q2.metric("Unique Farmers Synced", unique_farmers_synced)
    q3.metric("Synced Today", synced_today)
    q4.metric("Failed Syncs", field_failed_count)
    st.metric("Last Device Receipt", last_received)

    if not receipts_df.empty:
        st.dataframe(receipts_df.head(50), width="stretch")
    else:
        st.info("No AGROW Field device sync receipts have reached the central database yet.")

    st.markdown("#### Agent / Enumerator Performance")
    st.caption(
        "Registration counts come from the authoritative Farmer Register. "
        "Sync metrics come from the device receipt audit trail, so repeated transmissions do not inflate unique farmer counts."
    )
    agent_perf = _build_agent_performance(fresh_df, receipts_df)
    if not agent_perf.empty:
        st.dataframe(agent_perf, width="stretch")
    else:
        st.info("No agent performance data is available yet.")

    st.markdown("#### Field Operations Summary")
    st.caption(
        "Pitch view for the signed-in field user. Registration figures come from the Farmer Register; "
        "mobile synchronization figures come from the AGROW Field device receipt trail. "
        "The Streamlit fallback queue remains a backend diagnostic and is intentionally hidden here."
    )

    current_agent_registered = 0
    current_agent_registered_today = 0
    current_agent_unique_synced = 0
    current_agent_synced_today = 0
    current_agent_failed = 0
    current_agent_last_activity = "—"

    if not agent_perf.empty and "agent_id" in agent_perf.columns:
        current_agent_row = agent_perf[
            agent_perf["agent_id"].astype(str).str.upper() == str(user_id).upper()
        ]
        if not current_agent_row.empty:
            row = current_agent_row.iloc[0]
            current_agent_registered = int(row.get("registered_farmers", 0) or 0)
            current_agent_registered_today = int(row.get("registered_today", 0) or 0)
            current_agent_unique_synced = int(row.get("unique_farmers_synced", 0) or 0)
            current_agent_synced_today = int(row.get("synced_today", 0) or 0)
            current_agent_failed = int(row.get("failed_syncs", 0) or 0)
            last_activity_value = row.get("last_activity", pd.NaT)
            if pd.notna(last_activity_value):
                current_agent_last_activity = str(last_activity_value)

    f1, f2, f3 = st.columns(3)
    f1.metric("My Registered Farmers", current_agent_registered)
    f2.metric("My Registered Today", current_agent_registered_today)
    f3.metric("My Unique Mobile Syncs", current_agent_unique_synced)
    f4, f5, f6 = st.columns(3)
    f4.metric("My Syncs Today", current_agent_synced_today)
    f5.metric("My Failed Syncs", current_agent_failed)
    f6.metric("My Last Mobile Activity", current_agent_last_activity)

    if current_agent_failed == 0:
        st.success("AGROW Field mobile sync health: no failed syncs recorded for this field user.")
    else:
        st.warning(f"AGROW Field mobile sync health: {current_agent_failed} failed sync(s) require review.")


def generate_farmer_qr_code(farmer_id: str):
    verification_url = f"{get_effective_app_base_url()}?farmer_id={farmer_id}"

    qr = qrcode.make(verification_url)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)

    return buffer

def generate_farmer_id_card_pdf(selected_row, photo_path, logo_path=None):
    """Create a print-ready CR80 AGROW Farmer ID card with a signed QR code."""
    buffer = BytesIO()
    card_width = 85.6 * mm
    card_height = 54 * mm
    c = canvas.Canvas(buffer, pagesize=(card_width, card_height))

    def value(*keys, default="-"):
        for key in keys:
            raw = selected_row.get(key)
            if raw is not None and str(raw).strip() and str(raw).lower() != "nan":
                return str(raw).strip()
        return default

    def clipped(text, limit):
        text = str(text or "-")
        return text if len(text) <= limit else text[: max(1, limit - 1)] + "…"

    farmer_id = value("farmer_id", "Farmer_ID")
    farmer_name = value("farmer_full_name", "Farmer_Full_Name")
    state = normalize_state_name(value("state", "State"))
    lga = value("lga", "LGA")
    phone = value("phone_number", "Phone_Number")
    crop = value("primary_crop", "Primary_Crop")
    issue_date = datetime.now().strftime("%d %b %Y")

    # Background and border.
    c.setFillColorRGB(1, 1, 1)
    c.rect(0, 0, card_width, card_height, stroke=0, fill=1)
    c.setStrokeColorRGB(0.10, 0.36, 0.16)
    c.setLineWidth(1.2)
    c.roundRect(1.5, 1.5, card_width - 3, card_height - 3, 4, stroke=1, fill=0)

    # Branded header.
    header_h = 11 * mm
    c.setFillColorRGB(0.0, 0.29, 0.53)
    c.roundRect(1.5, card_height - header_h - 1.5, card_width - 3, header_h, 4, stroke=0, fill=1)
    c.rect(1.5, card_height - header_h - 1.5, card_width - 3, 4, stroke=0, fill=1)

    if logo_path and os.path.exists(logo_path):
        try:
            c.drawImage(
                logo_path,
                4 * mm,
                card_height - 9.7 * mm,
                width=8 * mm,
                height=8 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 9.2)
    c.drawString(14 * mm, card_height - 6.2 * mm, "AGROW FARMER ID CARD")
    c.setFont("Helvetica", 4.9)
    c.drawString(14 * mm, card_height - 8.7 * mm, "Agricultural Geographic Registration & Operations Workspace")

    # Farmer photograph.
    photo_x, photo_y = 4 * mm, 15 * mm
    photo_w, photo_h = 22 * mm, 23 * mm
    c.setFillColorRGB(0.96, 0.97, 0.96)
    c.setStrokeColorRGB(0.60, 0.68, 0.61)
    c.roundRect(photo_x, photo_y, photo_w, photo_h, 2, stroke=1, fill=1)
    if photo_path and os.path.exists(photo_path):
        try:
            c.drawImage(
                photo_path,
                photo_x + 1,
                photo_y + 1,
                width=photo_w - 2,
                height=photo_h - 2,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
        except Exception:
            c.setFillColorRGB(0.35, 0.35, 0.35)
            c.setFont("Helvetica", 6)
            c.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2, "PHOTO")
    else:
        c.setFillColorRGB(0.35, 0.35, 0.35)
        c.setFont("Helvetica", 6)
        c.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2, "PHOTO")

    # Farmer details.
    label_x = 29 * mm
    value_x = 41 * mm
    y = 37 * mm
    rows = [
        ("NAME", clipped(farmer_name.upper(), 26)),
        ("FARMER ID", clipped(farmer_id, 22)),
        ("STATE", clipped(state, 18)),
        ("LGA", clipped(lga, 18)),
        ("PHONE", clipped(phone, 18)),
        ("CROP", clipped(crop, 16)),
    ]
    for label, text in rows:
        c.setFillColorRGB(0.30, 0.34, 0.31)
        c.setFont("Helvetica-Bold", 4.8)
        c.drawString(label_x, y, label)
        c.setFillColorRGB(0.05, 0.08, 0.06)
        c.setFont("Helvetica-Bold" if label in {"NAME", "FARMER ID"} else "Helvetica", 5.8)
        c.drawString(value_x, y, text)
        y -= 4.1 * mm

    # Signed verification QR.
    signature = generate_qr_signature(farmer_id)
    verification_url = f"{get_effective_app_base_url()}?farmer_id={farmer_id}&sig={signature}"
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
    qr.add_data(verification_url)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_image.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    qr_size = 18 * mm
    qr_x = card_width - qr_size - 4 * mm
    qr_y = 4.8 * mm
    c.setFillColorRGB(1, 1, 1)
    c.setStrokeColorRGB(0.75, 0.75, 0.75)
    c.roundRect(qr_x - 1, qr_y - 1, qr_size + 2, qr_size + 4.2 * mm, 2, stroke=1, fill=1)
    c.drawImage(ImageReader(qr_buffer), qr_x, qr_y + 3.2 * mm, width=qr_size, height=qr_size, mask="auto")
    c.setFillColorRGB(0.10, 0.36, 0.16)
    c.setFont("Helvetica-Bold", 4.7)
    c.drawCentredString(qr_x + qr_size / 2, qr_y + 1.3 * mm, "SCAN TO VERIFY")

    # Footer and validation notice.
    c.setFillColorRGB(0.10, 0.36, 0.16)
    c.setFont("Helvetica-Bold", 5.1)
    c.drawString(4 * mm, 10.5 * mm, "VERIFIED AGRICULTURAL PROFILE")
    c.setFillColorRGB(0.25, 0.25, 0.25)
    c.setFont("Helvetica", 4.5)
    c.drawString(4 * mm, 7.8 * mm, f"Issued: {issue_date}")
    c.drawString(4 * mm, 5.5 * mm, "Not a national identity document. Validate through the signed QR code.")
    c.setFont("Helvetica-Oblique", 4.4)
    c.drawString(4 * mm, 3.3 * mm, "Issued through AGROW | Powered by DataDev Limited")

    c.save()
    buffer.seek(0)
    return buffer

# =========================================================
# 7. SESSION STATE INIT
# =========================================================
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if "user_id" not in st.session_state:
    st.session_state["user_id"] = None

if "role" not in st.session_state:
    st.session_state["role"] = None

if "farmer_form_version" not in st.session_state:
    st.session_state["farmer_form_version"] = 0

if "post_signup_redirect" not in st.session_state:
    st.session_state["post_signup_redirect"] = False

if "last_created_agent_id" not in st.session_state:
    st.session_state["last_created_agent_id"] = ""

# online / offline + sync monitor
if "db_online" not in st.session_state:
    st.session_state["db_online"] = True

if "last_auto_sync" not in st.session_state:
    st.session_state["last_auto_sync"] = None

if "last_sync_message" not in st.session_state:
    st.session_state["last_sync_message"] = ""

if "last_sync_status" not in st.session_state:
    st.session_state["last_sync_status"] = ""

if "auto_sync_ran" not in st.session_state:
    st.session_state["auto_sync_ran"] = False
# =========================================================
# 8. AUTH SCREEN
# =========================================================

# =========================================================
# MARKET PRICE INTELLIGENCE PERSISTENCE
# =========================================================
def ensure_market_price_table():
    """Create the field market-price table without changing existing AGROW tables."""
    id_column = "INTEGER PRIMARY KEY AUTOINCREMENT" if DB_BACKEND == "sqlite" else "BIGSERIAL PRIMARY KEY"
    ddl = f"""
    CREATE TABLE IF NOT EXISTS market_price_reports (
        id {id_column},
        commodity TEXT NOT NULL,
        product TEXT,
        market_name TEXT NOT NULL,
        lga TEXT,
        state TEXT NOT NULL,
        unit TEXT NOT NULL,
        farmgate_price REAL,
        wholesale_price REAL,
        retail_price REAL,
        min_price REAL,
        max_price REAL,
        typical_price REAL NOT NULL,
        source_type TEXT NOT NULL,
        source_name TEXT,
        notes TEXT,
        reported_by TEXT,
        reported_at TEXT NOT NULL,
        verified_at TEXT,
        status TEXT DEFAULT 'VERIFIED'
    )
    """
    with get_engine().begin() as conn:
        conn.execute(text(ddl))


def save_market_price_report(**payload):
    ensure_market_price_table()
    payload = dict(payload)
    payload.setdefault("reported_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    payload.setdefault("verified_at", payload["reported_at"])
    payload.setdefault("status", "VERIFIED")
    sql = text("""
        INSERT INTO market_price_reports (
            commodity, product, market_name, lga, state, unit,
            farmgate_price, wholesale_price, retail_price,
            min_price, max_price, typical_price,
            source_type, source_name, notes, reported_by,
            reported_at, verified_at, status
        ) VALUES (
            :commodity, :product, :market_name, :lga, :state, :unit,
            :farmgate_price, :wholesale_price, :retail_price,
            :min_price, :max_price, :typical_price,
            :source_type, :source_name, :notes, :reported_by,
            :reported_at, :verified_at, :status
        )
    """)
    with get_engine().begin() as conn:
        conn.execute(sql, payload)


def fetch_market_price_reports(limit=250):
    ensure_market_price_table()
    try:
        with get_engine().connect() as conn:
            rows = conn.execute(text("""
                SELECT * FROM market_price_reports
                WHERE COALESCE(status, 'VERIFIED') = 'VERIFIED'
                ORDER BY reported_at DESC, id DESC
                LIMIT :limit
            """), {"limit": int(limit)}).mappings().all()
        return pd.DataFrame([dict(r) for r in rows])
    except Exception:
        return pd.DataFrame()


def _public_listing_url(listing_id) -> str:
    return f"{get_effective_app_base_url()}?market=1&listing_id={int(listing_id)}"


def render_public_listing_detail(listing_id):
    """Render one public, shareable produce listing for consumers/off-takers."""
    listings = fetch_market_listings(status="AVAILABLE")
    if listings.empty:
        st.warning("This MarketLink listing is no longer available.")
        return
    try:
        target_id = int(listing_id)
    except Exception:
        st.error("Invalid MarketLink listing reference.")
        return
    record = listings[listings["id"].astype(int) == target_id]
    if record.empty:
        st.warning("This MarketLink listing is unavailable, sold, paused or withdrawn.")
        return
    item = record.iloc[0]
    st.markdown("## 🌾 AGROW MarketLink Listing")
    st.success("✅ Produce listing from a verified AGROW farmer" if str(item.get("nin_status", "")).lower() == "verified" else "Produce listing from a registered AGROW farmer")
    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown(f"### {item.get('product', '-')}")
        st.write(f"**Commodity:** {item.get('commodity', '-')}")
        st.write(f"**Available:** {float(item.get('available_quantity') or 0):g} {item.get('unit', '')}")
        st.write(f"**Asking price:** {_money(item.get('price'))} / {item.get('unit', '')}")
        st.write(f"**Ready / harvest date:** {item.get('ready_date') or '-'}")
        st.write(f"**Location:** {item.get('community') or '-'}, {item.get('lga') or '-'}, {normalize_state_name(item.get('state') or '-')}")
        if item.get('description'):
            st.caption(str(item.get('description')))
    with c2:
        st.write(f"**Seller:** {item.get('farmer_full_name') or '-'}")
        st.write(f"**Farmer ID:** {item.get('farmer_id') or '-'}")
        seller_phone = str(item.get('phone_number') or '')
        if seller_phone:
            st.write(f"**Phone:** {seller_phone}")
            wa_phone = _phone_for_whatsapp(seller_phone)
            if wa_phone:
                message = quote(_seller_message(item))
                st.link_button("💬 WhatsApp Seller", f"https://wa.me/{wa_phone}?text={message}", width="stretch", type="primary")
            st.link_button("📞 Call Seller", f"tel:{seller_phone}", width="stretch")
        farmer_id = str(item.get('farmer_id') or '').strip()
        if farmer_id:
            verify_url = f"{get_effective_app_base_url()}?farmer_id={farmer_id}&sig={generate_qr_signature(farmer_id)}"
            st.link_button("✅ Verify Farmer", verify_url, width="stretch")
    with st.expander("📨 Request Quote / Send Buyer Enquiry", expanded=True):
        with st.form(f"public_listing_enquiry_{target_id}"):
            buyer_name = st.text_input("Your name")
            buyer_phone = st.text_input("Phone number")
            qty = st.number_input("Quantity requested", min_value=0.0, step=1.0)
            enquiry_message = st.text_area("Message", placeholder="Delivery location, preferred date, quality/grade or other requirements")
            if st.form_submit_button("Send Enquiry", type="primary", use_container_width=True):
                if not buyer_name.strip() or not buyer_phone.strip():
                    st.error("Your name and phone number are required.")
                else:
                    create_market_enquiry(target_id, buyer_name.strip(), buyer_phone.strip(), qty, enquiry_message.strip())
                    st.success("Enquiry sent to AGROW MarketLink.")
    st.caption("Share this page directly with consumers, processors, retailers and off-takers.")
    st.code(_public_listing_url(target_id), language=None)

# MARKETLINK — UNIVERSAL AGRICULTURAL MARKETPLACE
# =========================================================
MARKET_COMMODITIES = [
    "Poultry", "Maize", "Rice", "Soybean", "Cassava", "Yam", "Sorghum",
    "Millet", "Groundnut", "Cowpea", "Wheat", "Tomato", "Pepper", "Onion",
    "Vegetables", "Fish", "Cattle", "Goat", "Sheep", "Pig", "Eggs", "Other",
]

MARKET_UNITS = [
    "kg", "tonne", "bag", "crate", "basket", "bird", "tray", "litre",
    "bundle", "tuber", "head", "unit",
]

INPUT_CATEGORIES = [
    "Seeds & Planting Materials", "Fertilizer", "Agrochemicals", "Animal Feed",
    "Veterinary", "Farm Machinery", "Irrigation", "Storage", "Packaging", "Other",
]


def _phone_for_whatsapp(phone: str) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    if digits.startswith("0") and len(digits) == 11:
        return "234" + digits[1:]
    return digits


def _money(value) -> str:
    try:
        return f"₦{float(value):,.2f}"
    except Exception:
        return "-"


def _seller_message(item) -> str:
    product = str(item.get("product") or "produce")
    commodity = str(item.get("commodity") or "")
    farmer = str(item.get("farmer_full_name") or "farmer")
    quantity = float(item.get("available_quantity") or 0)
    unit = str(item.get("unit") or "unit")
    return (
        f"Hello {farmer}, I found your {product} listing on AGROW MarketLink. "
        f"Commodity: {commodity}. Available quantity: {quantity:g} {unit}. "
        "Please confirm current availability, price and delivery/collection options. Thank you."
    )


def _render_seller_contact_panel(item, key_prefix: str) -> None:
    """Render robust seller-contact options without changing farmer registry data."""
    phone = str(item.get("phone_number") or "").strip()
    wa_phone = _phone_for_whatsapp(phone)
    farmer_id = str(item.get("farmer_id") or "").strip()
    message = quote(_seller_message(item), safe="")

    if wa_phone:
        c1, c2 = st.columns(2)
        c1.link_button(
            "💬 WhatsApp",
            f"https://wa.me/{wa_phone}?text={message}",
            width="stretch",
        )
        c2.link_button(
            "🖥️ WhatsApp Web",
            f"https://web.whatsapp.com/send?phone={wa_phone}&text={message}",
            width="stretch",
        )

    if phone:
        safe_tel = "+" + wa_phone if wa_phone else re.sub(r"[^0-9+]", "", phone)
        st.markdown(
            f'<a href="tel:{safe_tel}" style="display:block;text-align:center;padding:0.55rem 0.8rem;'
            'border:1px solid #d0d7de;border-radius:0.5rem;text-decoration:none;font-weight:600;'
            'margin:0.25rem 0;">📞 Call Seller</a>',
            unsafe_allow_html=True,
        )
        st.caption(f"Seller phone: {phone}")

    latitude = item.get("latitude")
    longitude = item.get("longitude")
    try:
        lat = float(latitude)
        lon = float(longitude)
    except (TypeError, ValueError):
        lat = lon = 0.0
    if abs(lat) > 0.000001 or abs(lon) > 0.000001:
        st.link_button(
            "📍 View Farm Location",
            f"https://www.google.com/maps/search/?api=1&query={lat},{lon}",
            width="stretch",
        )

    with st.expander("👨‍🌾 View Farmer Profile"):
        pc1, pc2 = st.columns([1, 2])
        photo_path = resolve_farmer_photo_path(item.get("photo_path", ""), farmer_id)
        if photo_path:
            pc1.image(photo_path, width=180)
        else:
            pc1.info("Farmer photo not available on this server.")
        verified = str(item.get("nin_status") or "").lower() == "verified"
        pc2.markdown(f"**{item.get('farmer_full_name') or '-'}**")
        pc2.write(f"Farmer ID: **{farmer_id or '-'}**")
        pc2.write(f"Verification: **{'Verified' if verified else 'Registered'}**")
        pc2.write(f"Primary crop: **{item.get('primary_crop') or '-'}**")
        pc2.write(
            f"Location: **{item.get('community') or '-'}, {item.get('lga') or '-'}, "
            f"{normalize_state_name(item.get('state') or '-')}**"
        )
        if farmer_id:
            verification_url = (
                f"{get_effective_app_base_url()}?farmer_id={farmer_id}"
                f"&sig={generate_qr_signature(farmer_id)}"
            )
            pc2.link_button("✅ Open Verified Farmer Record", verification_url, width="stretch")


def render_public_marketlink():
    """Public MarketLink browser for consumers, buyers and input discovery."""
    st.markdown("### 🛒 AGROW MarketLink")
    st.caption(
        "Discover produce from verified AGROW farmers and agricultural inputs from listed suppliers. "
        "Poultry is the first live demonstration; the same marketplace supports every agricultural commodity."
    )

    produce_tab, inputs_tab, prices_tab = st.tabs(
        ["🌾 Buy Produce", "🧰 Agricultural Inputs", "📊 Market Price Board"]
    )

    with produce_tab:
        listings = fetch_market_listings(status="AVAILABLE")
        if listings.empty:
            st.info("No produce listings are available yet. Verified farmer listings will appear here.")
        else:
            filter1, filter2, filter3 = st.columns(3)
            commodity_options = ["All"] + sorted(listings["commodity"].dropna().unique().tolist())
            state_options = ["All"] + sorted(listings["state"].dropna().unique().tolist())
            selected_commodity = filter1.selectbox("Commodity", commodity_options, key="public_market_commodity")
            selected_state = filter2.selectbox("State", state_options, key="public_market_state")
            query = filter3.text_input("Search product/location", key="public_market_search")

            shown = listings.copy()
            if selected_commodity != "All":
                shown = shown[shown["commodity"] == selected_commodity]
            if selected_state != "All":
                shown = shown[shown["state"] == selected_state]
            if query.strip():
                q = query.strip().lower()
                mask = (
                    shown["product"].fillna("").str.lower().str.contains(q, regex=False)
                    | shown["community"].fillna("").str.lower().str.contains(q, regex=False)
                    | shown["lga"].fillna("").str.lower().str.contains(q, regex=False)
                )
                shown = shown[mask]

            st.caption(f"{len(shown)} available listing(s)")
            for _, item in shown.head(30).iterrows():
                with st.container(border=True):
                    c1, c2 = st.columns([3, 1])
                    verified = str(item.get("nin_status", "")).lower() == "verified"
                    badge = "✅ Verified Farmer" if verified else "Registered Farmer"
                    c1.markdown(f"#### {item.get('product', '-')}")
                    c1.caption(f"{badge} · {item.get('farmer_full_name', '-')}")
                    c1.write(
                        f"**{item.get('commodity', '-')}** · {float(item.get('available_quantity') or 0):g} "
                        f"{item.get('unit', '')} available · **{_money(item.get('price'))} / {item.get('unit', '')}**"
                    )
                    c1.write(
                        f"📍 {item.get('community') or '-'}, {item.get('lga') or '-'}, "
                        f"{normalize_state_name(item.get('state') or '-')}"
                    )
                    if item.get("ready_date"):
                        c1.caption(f"Ready / harvest date: {item.get('ready_date')}")
                    if item.get("description"):
                        c1.caption(str(item.get("description")))

                    farmer_id = str(item.get("farmer_id") or "")
                    if farmer_id:
                        verification_url = (
                            f"{get_effective_app_base_url()}?farmer_id={farmer_id}"
                            f"&sig={generate_qr_signature(farmer_id)}"
                        )
                        c2.link_button("✅ Verify Farmer", verification_url, width="stretch")
                    c2.link_button("🌐 Open / Share Listing", _public_listing_url(item["id"]), width="stretch", type="primary")
                    c2.caption(f"Farmer ID: {farmer_id or '-'}")

                    with st.expander("📞 Contact Seller", expanded=False):
                        _render_seller_contact_panel(item, f"seller_{int(item['id'])}")

                    with st.expander("📨 Request Quote / Send Buyer Enquiry"):
                        with st.form(f"public_enquiry_{int(item['id'])}"):
                            buyer_name = st.text_input("Your name")
                            buyer_phone = st.text_input("Phone number")
                            buyer_org = st.text_input("Organisation / Business (optional)")
                            qty = st.number_input(
                                f"Quantity required ({item.get('unit', 'unit')})",
                                min_value=0.0,
                                step=1.0,
                            )
                            preferred_date = st.date_input("Preferred delivery / collection date")
                            message = st.text_area("Message / delivery requirement")
                            if st.form_submit_button("Send Enquiry", use_container_width=True):
                                if not buyer_name.strip() or not buyer_phone.strip():
                                    st.error("Name and phone number are required.")
                                elif qty <= 0:
                                    st.error("Enter the quantity required.")
                                else:
                                    enquiry_message = (
                                        f"Organisation: {buyer_org.strip() or '-'} | "
                                        f"Preferred date: {preferred_date} | {message.strip()}"
                                    )
                                    create_market_enquiry(
                                        int(item["id"]), buyer_name.strip(), buyer_phone.strip(), qty, enquiry_message
                                    )
                                    st.success("Enquiry recorded in AGROW. Use Contact Seller if you also want to call or message the farmer directly.")

    with inputs_tab:
        products = fetch_input_products(status="AVAILABLE")
        if products.empty:
            st.info("No agricultural input listings are available yet.")
        else:
            f1, f2 = st.columns(2)
            categories = ["All"] + sorted(products["category"].dropna().unique().tolist())
            category = f1.selectbox("Input category", categories, key="public_input_category")
            input_query = f2.text_input("Search product / commodity", key="public_input_search")
            shown_inputs = products.copy()
            if category != "All":
                shown_inputs = shown_inputs[shown_inputs["category"] == category]
            if input_query.strip():
                q = input_query.strip().lower()
                shown_inputs = shown_inputs[
                    shown_inputs["product_name"].fillna("").str.lower().str.contains(q, regex=False)
                    | shown_inputs["applicable_commodities"].fillna("").str.lower().str.contains(q, regex=False)
                ]
            for _, item in shown_inputs.head(30).iterrows():
                with st.container(border=True):
                    a, b = st.columns([3, 1])
                    a.markdown(f"#### {item.get('product_name', '-')}")
                    a.caption(f"{item.get('category', '-')} · Supplier: {item.get('supplier_name', '-')}")
                    a.write(
                        f"**{_money(item.get('price'))} / {item.get('unit', '')}** · "
                        f"{float(item.get('quantity') or 0):g} {item.get('unit', '')} listed"
                    )
                    a.write(f"For: {item.get('applicable_commodities') or 'Multiple commodities'}")
                    a.caption(f"📍 {item.get('lga') or '-'}, {normalize_state_name(item.get('state') or '-')}")
                    if item.get("description"):
                        a.caption(str(item.get("description")))
                    phone = str(item.get("supplier_phone") or "")
                    wa_phone = _phone_for_whatsapp(phone)
                    if wa_phone:
                        b.link_button(
                            "💬 Contact Supplier",
                            f"https://wa.me/{wa_phone}?text=Hello%2C%20I%20found%20your%20agricultural%20input%20listing%20on%20AGROW%20MarketLink.",
                            width="stretch",
                        )
                    if phone:
                        b.caption(f"☎ {phone}")

    with prices_tab:
        st.markdown("#### 📊 Current Market Price Intelligence")
        st.caption("AGROW separates seller asking prices from field-reported market prices so users can see the source of every figure.")

        reports = fetch_market_price_reports()
        if not reports.empty:
            latest = reports.copy()
            for col in ["typical_price", "min_price", "max_price", "farmgate_price", "wholesale_price", "retail_price"]:
                if col in latest.columns:
                    latest[col] = pd.to_numeric(latest[col], errors="coerce")
            display_cols = [
                "reported_at", "commodity", "product", "market_name", "lga", "state", "unit",
                "typical_price", "min_price", "max_price", "source_type", "source_name"
            ]
            existing_cols = [c for c in display_cols if c in latest.columns]
            st.dataframe(latest[existing_cols].head(60), width="stretch", hide_index=True)
            st.caption("Field Market Price = time-stamped price captured by an authorised AGROW source at a named market.")
        else:
            st.info("No verified field market-price reports have been submitted yet.")

        listings = fetch_market_listings(status="AVAILABLE")
        st.markdown("#### 🌾 AGROW Seller Listing Prices")
        if listings.empty:
            st.info("Seller price statistics will populate automatically as produce listings are published.")
        else:
            board = (
                listings.groupby(["commodity", "product", "state", "unit"], dropna=False)
                .agg(
                    Listings=("id", "count"),
                    Available=("available_quantity", "sum"),
                    **{"Min Price": ("price", "min"), "Average Price": ("price", "mean"), "Max Price": ("price", "max")},
                )
                .reset_index()
            )
            for col in ["Min Price", "Average Price", "Max Price"]:
                board[col] = board[col].map(lambda v: f"₦{float(v):,.2f}")
            st.dataframe(
                board[["commodity", "product", "state", "unit", "Listings", "Available", "Min Price", "Average Price", "Max Price"]],
                width="stretch", hide_index=True,
            )
            st.caption("AGROW Seller Listing Price = current asking prices from active MarketLink listings; it is not the same as a completed trade or official exchange benchmark.")


def render_marketlink_workspace(df: pd.DataFrame, role: str, user_id: str):
    st.markdown("## 🛒 AGROW MarketLink")
    st.caption(
        "Universal agricultural marketplace connecting verified farmers to consumers, off-takers and input suppliers. "
        "Poultry is the first demonstration commodity; no additional feature build is required for maize, rice, soybean or other commodities."
    )

    listings = fetch_market_listings()
    available = listings[listings["status"] == "AVAILABLE"] if not listings.empty else listings
    input_products = fetch_input_products()
    available_inputs = input_products[input_products["status"] == "AVAILABLE"] if not input_products.empty else input_products

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Available Produce Listings", len(available))
    m2.metric("Commodities Listed", available["commodity"].nunique() if not available.empty else 0)
    m3.metric("Input Products", len(available_inputs))
    m4.metric("Buyer Enquiries", len(fetch_market_enquiries()))

    sell_tab, browse_tab, input_tab, price_report_tab, manage_tab = st.tabs(
        [
            "➕ Create Produce Listing",
            "🌾 Browse Market",
            "🧰 Input Marketplace",
            "📍 Report Market Price",
            "📋 Manage",
        ]
    )

    with sell_tab:
        eligible = df.copy()
        if role == "agent" and not eligible.empty:
            eligible = eligible[eligible["agent_id"] == user_id]
        if not eligible.empty:
            eligible = eligible[eligible["nin_status"].astype(str).str.lower() == "verified"]

        if eligible.empty:
            st.warning("No verified farmer is available to create a market listing. Verify a farmer record first.")
        else:
            eligible = eligible.copy()
            eligible["market_label"] = eligible.apply(
                lambda r: f"{r['farmer_id']} — {r['farmer_full_name']} ({r['state']} / {r['lga']})", axis=1
            )
            with st.form("create_market_listing_form", clear_on_submit=True):
                farmer_label = st.selectbox("Verified Farmer", eligible["market_label"].tolist())
                farmer = eligible[eligible["market_label"] == farmer_label].iloc[0]
                c1, c2 = st.columns(2)
                commodity = c1.selectbox("Commodity", MARKET_COMMODITIES)
                product = c2.text_input("Product / Variety", placeholder="e.g. Broiler Chicken, Paddy Rice, Yellow Maize")
                q1, q2, q3 = st.columns(3)
                quantity = q1.number_input("Quantity Available", min_value=0.0, step=1.0)
                unit = q2.selectbox("Unit", MARKET_UNITS)
                price = q3.number_input("Asking Price per Unit (₦)", min_value=0.0, step=100.0)
                ready_date = st.date_input("Ready / Harvest Date")
                description = st.text_area("Listing Description", placeholder="Quality, grade, average weight, packaging, collection/delivery notes...")
                st.caption(
                    f"Location will be inherited from the farmer record: {farmer.get('community_village') or '-'}, "
                    f"{farmer.get('lga') or '-'}, {normalize_state_name(farmer.get('state') or '-')}"
                )
                submit = st.form_submit_button("Publish to MarketLink", use_container_width=True, type="primary")
                if submit:
                    if not product.strip():
                        st.error("Product / variety is required.")
                    elif quantity <= 0:
                        st.error("Quantity must be greater than zero.")
                    elif price <= 0:
                        st.error("Price must be greater than zero.")
                    else:
                        create_market_listing(
                            farmer_id=str(farmer["farmer_id"]),
                            commodity=commodity,
                            product=product.strip(),
                            quantity=quantity,
                            unit=unit,
                            price=price,
                            ready_date=str(ready_date),
                            state=str(farmer.get("state") or ""),
                            lga=str(farmer.get("lga") or ""),
                            community=str(farmer.get("community_village") or ""),
                            description=description.strip(),
                            created_by=user_id,
                        )
                        log_action(user_id, "MARKET_LISTING_CREATED", f"{farmer['farmer_id']} | {commodity} | {product}")
                        st.success("Produce listing published to AGROW MarketLink.")
                        st.rerun()

    with browse_tab:
        # Reuse the same public market experience inside the authenticated workspace.
        render_public_marketlink()

    with input_tab:
        st.markdown("### List Agricultural Inputs")
        st.caption("This generic input directory serves poultry, grains, legumes, roots, vegetables, livestock, aquaculture and future commodities.")
        with st.form("create_input_product_form", clear_on_submit=True):
            i1, i2 = st.columns(2)
            supplier_name = i1.text_input("Supplier / Business Name")
            supplier_phone = i2.text_input("Supplier Phone")
            i3, i4 = st.columns(2)
            category = i3.selectbox("Input Category", INPUT_CATEGORIES)
            product_name = i4.text_input("Product Name", placeholder="e.g. Broiler Starter Feed, Urea Fertilizer")
            applicable = st.multiselect("Applicable Commodities", MARKET_COMMODITIES, default=[])
            i5, i6, i7 = st.columns(3)
            input_qty = i5.number_input("Quantity Available", min_value=0.0, step=1.0)
            input_unit = i6.selectbox("Unit", MARKET_UNITS, key="input_market_unit")
            input_price = i7.number_input("Price per Unit (₦)", min_value=0.0, step=100.0)
            i8, i9 = st.columns(2)
            input_state = i8.selectbox("State", sorted(NIGERIA_LGA_MAP.keys()), key="input_market_state")
            input_lgas = NIGERIA_LGA_MAP.get(input_state, [])
            input_lga = i9.selectbox("LGA", input_lgas if input_lgas else ["N/A"], key="input_market_lga")
            input_description = st.text_area("Description / Specification")
            add_input = st.form_submit_button("Publish Input Product", use_container_width=True)
            if add_input:
                if not supplier_name.strip() or not supplier_phone.strip():
                    st.error("Supplier name and phone are required.")
                elif not product_name.strip():
                    st.error("Product name is required.")
                elif input_qty <= 0 or input_price <= 0:
                    st.error("Quantity and price must be greater than zero.")
                else:
                    create_input_product(
                        supplier_name=supplier_name.strip(), supplier_phone=supplier_phone.strip(),
                        category=category, product_name=product_name.strip(),
                        applicable_commodities=", ".join(applicable) if applicable else "All / General",
                        quantity=input_qty, unit=input_unit, price=input_price,
                        state=input_state, lga=input_lga, description=input_description.strip(),
                        created_by=user_id,
                    )
                    log_action(user_id, "INPUT_PRODUCT_CREATED", f"{supplier_name} | {product_name}")
                    st.success("Agricultural input published to MarketLink.")
                    st.rerun()

    with price_report_tab:
        st.markdown("### 📍 Report Current Market Price")
        st.caption("For authorised agents/administrators: capture a named market, unit, source and timestamp. This feeds the public Market Price Board separately from farmer asking prices.")
        with st.form("market_price_report_form", clear_on_submit=True):
            r1, r2, r3 = st.columns(3)
            rp_commodity = r1.selectbox("Commodity", MARKET_COMMODITIES, key="price_report_commodity")
            rp_product = r2.text_input("Product / Variety / Grade", placeholder="e.g. Yellow Maize, Paddy Rice, Broiler Chicken")
            rp_unit = r3.selectbox("Unit", MARKET_UNITS, key="price_report_unit")
            r4, r5, r6 = st.columns(3)
            rp_state = r4.selectbox("State", sorted(NIGERIA_LGA_MAP.keys()), key="price_report_state")
            rp_lgas = NIGERIA_LGA_MAP.get(rp_state, [])
            rp_lga = r5.selectbox("LGA", rp_lgas if rp_lgas else ["N/A"], key="price_report_lga")
            rp_market = r6.text_input("Market Name", placeholder="e.g. Abaji Main Market")
            p1, p2, p3 = st.columns(3)
            rp_min = p1.number_input("Observed Minimum (₦)", min_value=0.0, step=100.0)
            rp_typical = p2.number_input("Typical / Average (₦)", min_value=0.0, step=100.0)
            rp_max = p3.number_input("Observed Maximum (₦)", min_value=0.0, step=100.0)
            p4, p5, p6 = st.columns(3)
            rp_farmgate = p4.number_input("Farm-gate (₦, optional)", min_value=0.0, step=100.0)
            rp_wholesale = p5.number_input("Wholesale (₦, optional)", min_value=0.0, step=100.0)
            rp_retail = p6.number_input("Retail (₦, optional)", min_value=0.0, step=100.0)
            rp_source_type = st.selectbox("Source Type", ["AGROW Field Enumerator", "Market Association", "Government / Official", "Partner Dataset", "Other"])
            rp_source_name = st.text_input("Source / Contact / Reference", placeholder="e.g. Abaji Market Association / Daily field survey")
            rp_notes = st.text_area("Notes", placeholder="Quality, bag size, moisture, live weight, grade, transaction context, etc.")
            if st.form_submit_button("✅ Publish Verified Market Price", type="primary", use_container_width=True):
                if not rp_market.strip():
                    st.error("Market name is required.")
                elif rp_typical <= 0:
                    st.error("Typical / average market price must be greater than zero.")
                else:
                    save_market_price_report(
                        commodity=rp_commodity, product=rp_product.strip(), market_name=rp_market.strip(),
                        lga=rp_lga, state=rp_state, unit=rp_unit,
                        farmgate_price=rp_farmgate or None, wholesale_price=rp_wholesale or None, retail_price=rp_retail or None,
                        min_price=rp_min or None, max_price=rp_max or None, typical_price=rp_typical,
                        source_type=rp_source_type, source_name=rp_source_name.strip(), notes=rp_notes.strip(),
                        reported_by=user_id,
                    )
                    log_action(user_id, "MARKET_PRICE_REPORTED", f"{rp_commodity} | {rp_market} | {rp_typical}")
                    st.success("Market price published to the AGROW public Price Board.")

    with manage_tab:
        st.markdown("### Produce Listings")
        current = fetch_market_listings()
        if role == "agent" and not current.empty:
            # Agents manage only listings they created; administrators retain platform-wide oversight.
            current = current[current["created_by"] == user_id]
        if current.empty:
            st.info("No produce listings to manage yet.")
        else:
            st.dataframe(
                current[["id", "farmer_id", "commodity", "product", "available_quantity", "unit", "price", "status", "created_at"]],
                width="stretch", hide_index=True,
            )
            options = {
                f"#{int(r['id'])} — {r['farmer_id']} — {r['product']} — {r['status']}": int(r["id"])
                for _, r in current.iterrows()
            }
            selected = st.selectbox("Select listing to update", list(options.keys()), key="manage_listing_select")
            new_status = st.selectbox("Listing Status", ["AVAILABLE", "RESERVED", "SOLD", "WITHDRAWN"], key="manage_listing_status")
            if st.button("Update Listing Status", key="update_market_status"):
                update_listing_status(options[selected], new_status)
                log_action(user_id, "MARKET_LISTING_STATUS", f"{options[selected]} -> {new_status}")
                st.success("Listing status updated.")
                st.rerun()

        if role == "admin":
            st.markdown("### Buyer Enquiries")
            enquiries = fetch_market_enquiries()
            if enquiries.empty:
                st.info("No buyer enquiries yet.")
            else:
                st.dataframe(enquiries, width="stretch", hide_index=True)

            st.markdown("### Input Listings")
            products = fetch_input_products()
            if products.empty:
                st.info("No input products listed yet.")
            else:
                st.dataframe(products, width="stretch", hide_index=True)


# =========================================================

# =========================================================
# FARMER MARKET PORTAL — VERIFIED FARMER SELF-SERVICE
# =========================================================
def _normalise_registered_phone(value: str) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if digits.startswith("234") and len(digits) == 13:
        digits = "0" + digits[3:]
    return digits


def _farmer_record_by_id(farmer_id: str):
    farmer_id = str(farmer_id or "").strip().upper()
    if not farmer_id:
        return None
    farmers = fetch_farmers()
    if farmers.empty or "farmer_id" not in farmers.columns:
        return None
    match = farmers[
        farmers["farmer_id"].astype(str).str.strip().str.upper() == farmer_id
    ]
    return None if match.empty else match.iloc[0]


def _ensure_farmer_market_credentials_table() -> None:
    with get_engine().begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS farmer_market_credentials (
                farmer_id TEXT PRIMARY KEY,
                pin_hash TEXT NOT NULL,
                salt TEXT NOT NULL,
                activated_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """))


def _farmer_credential(farmer_id: str):
    _ensure_farmer_market_credentials_table()
    with get_engine().connect() as conn:
        row = conn.execute(
            text("SELECT farmer_id, pin_hash, salt, activated_at, updated_at FROM farmer_market_credentials WHERE farmer_id=:farmer_id"),
            {"farmer_id": str(farmer_id or "").strip().upper()},
        ).mappings().first()
    return row


def _hash_farmer_pin(farmer_id: str, pin: str, salt_hex: str) -> str:
    material = f"{str(farmer_id).strip().upper()}::{str(pin)}".encode("utf-8")
    return hashlib.pbkdf2_hmac(
        "sha256",
        material,
        bytes.fromhex(salt_hex),
        160_000,
    ).hex()


def _set_farmer_pin(farmer_id: str, pin: str) -> None:
    farmer_id = str(farmer_id or "").strip().upper()
    salt_hex = os.urandom(16).hex()
    pin_hash = _hash_farmer_pin(farmer_id, pin, salt_hex)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _ensure_farmer_market_credentials_table()
    with get_engine().begin() as conn:
        existing = conn.execute(
            text("SELECT farmer_id FROM farmer_market_credentials WHERE farmer_id=:farmer_id"),
            {"farmer_id": farmer_id},
        ).first()
        if existing:
            conn.execute(
                text("""
                    UPDATE farmer_market_credentials
                    SET pin_hash=:pin_hash, salt=:salt, updated_at=:updated_at
                    WHERE farmer_id=:farmer_id
                """),
                {
                    "farmer_id": farmer_id,
                    "pin_hash": pin_hash,
                    "salt": salt_hex,
                    "updated_at": now,
                },
            )
        else:
            conn.execute(
                text("""
                    INSERT INTO farmer_market_credentials
                    (farmer_id, pin_hash, salt, activated_at, updated_at)
                    VALUES (:farmer_id, :pin_hash, :salt, :activated_at, :updated_at)
                """),
                {
                    "farmer_id": farmer_id,
                    "pin_hash": pin_hash,
                    "salt": salt_hex,
                    "activated_at": now,
                    "updated_at": now,
                },
            )


def _verify_farmer_pin(farmer_id: str, pin: str) -> bool:
    cred = _farmer_credential(farmer_id)
    if not cred:
        return False
    candidate = _hash_farmer_pin(farmer_id, pin, str(cred["salt"]))
    return hmac.compare_digest(candidate, str(cred["pin_hash"]))


def _farmer_is_verified(farmer) -> bool:
    if farmer is None:
        return False
    return str(farmer.get("nin_status", "") or "").strip().lower() == "verified"


def render_farmer_market_login() -> None:
    st.markdown("### 🌾 Farmer Market Portal")
    st.caption(
        "Verified AGROW farmers can publish produce directly to MarketLink. "
        "Your Farmer ID is your username. On first activation, use the phone number registered during enumeration, then create a private PIN."
    )

    pending_farmer_id = st.session_state.get("farmer_activation_id", "")
    if pending_farmer_id:
        farmer = _farmer_record_by_id(pending_farmer_id)
        if farmer is None:
            st.session_state.pop("farmer_activation_id", None)
            st.error("Farmer record is no longer available. Start activation again.")
            return

        st.success(f"Identity confirmed for {farmer.get('farmer_full_name', 'Farmer')} — {pending_farmer_id}")
        p1, p2 = st.columns(2)
        new_pin = p1.text_input("Create 4–6 digit PIN", type="password", max_chars=6, key="farmer_new_pin")
        confirm_pin = p2.text_input("Confirm PIN", type="password", max_chars=6, key="farmer_confirm_pin")
        if st.button("🔐 Activate Farmer Market Access", width="stretch", type="primary", key="activate_farmer_market"):
            if not re.fullmatch(r"\d{4,6}", str(new_pin or "")):
                st.error("PIN must contain 4 to 6 digits.")
            elif new_pin != confirm_pin:
                st.error("PIN entries do not match.")
            else:
                _set_farmer_pin(pending_farmer_id, new_pin)
                st.session_state.pop("farmer_activation_id", None)
                st.session_state.logged_in = True
                st.session_state.user_id = pending_farmer_id
                st.session_state.role = "farmer"
                log_action(pending_farmer_id, "FARMER_MARKET_ACTIVATED", "Farmer Market Portal PIN created")
                st.rerun()
        if st.button("Cancel activation", width="stretch", key="cancel_farmer_activation"):
            st.session_state.pop("farmer_activation_id", None)
            st.rerun()
        return

    farmer_id = st.text_input(
        "Farmer ID",
        placeholder="e.g. AG-20260808111836",
        key="farmer_market_login_id",
    ).strip().upper()
    access_code = st.text_input(
        "PIN / Registered Phone Number",
        type="password",
        key="farmer_market_login_code",
    )

    if st.button("🌾 Open Farmer Market Portal", width="stretch", type="primary", key="farmer_market_login_button"):
        farmer = _farmer_record_by_id(farmer_id)
        if farmer is None:
            st.error("Farmer ID was not found in the active AGROW registry.")
            return
        if not _farmer_is_verified(farmer):
            st.error("This Farmer ID is not yet verified. Farmer Market access is available only to verified records.")
            return

        cred = _farmer_credential(farmer_id)
        if cred:
            if not _verify_farmer_pin(farmer_id, access_code):
                st.error("Invalid Farmer ID or PIN.")
                return
            st.session_state.logged_in = True
            st.session_state.user_id = farmer_id
            st.session_state.role = "farmer"
            log_action(farmer_id, "FARMER_MARKET_LOGIN", "Successful farmer market login")
            st.rerun()
        else:
            registered_phone = _normalise_registered_phone(farmer.get("phone_number", ""))
            supplied_phone = _normalise_registered_phone(access_code)
            if not registered_phone or supplied_phone != registered_phone:
                st.error("First-time activation requires the phone number registered on this Farmer ID.")
                return
            st.session_state["farmer_activation_id"] = farmer_id
            st.rerun()

    with st.expander("Forgot PIN / Reset Access"):
        st.caption("Pilot reset: confirm the Farmer ID and registered phone number, then create a new PIN.")
        reset_id = st.text_input("Farmer ID for reset", key="farmer_reset_id").strip().upper()
        reset_phone = st.text_input("Registered phone number", type="password", key="farmer_reset_phone")
        if st.button("Verify and Reset PIN", width="stretch", key="farmer_reset_button"):
            farmer = _farmer_record_by_id(reset_id)
            if farmer is None or not _farmer_is_verified(farmer):
                st.error("Verified farmer record not found.")
            elif _normalise_registered_phone(reset_phone) != _normalise_registered_phone(farmer.get("phone_number", "")):
                st.error("Registered phone number does not match this Farmer ID.")
            else:
                st.session_state["farmer_activation_id"] = reset_id
                st.rerun()


def render_farmer_market_portal(farmer_id: str) -> None:
    farmer = _farmer_record_by_id(farmer_id)
    if farmer is None:
        st.error("Your farmer record could not be loaded. Please sign in again.")
        st.session_state.logged_in = False
        st.session_state.user_id = None
        st.session_state.role = None
        return

    if not _farmer_is_verified(farmer):
        st.error("This farmer record is not currently verified for MarketLink self-service.")
        return

    farmer_id = str(farmer.get("farmer_id", "") or "").strip().upper()
    farmer_name = str(farmer.get("farmer_full_name", "Farmer") or "Farmer")
    photo_path = resolve_farmer_photo_path(farmer.get("photo_path", ""), farmer_id)

    st.sidebar.markdown("### 🌾 Farmer Market")
    st.sidebar.write(f"**{farmer_name}**")
    st.sidebar.caption(farmer_id)
    st.sidebar.write("✅ Verified Farmer")
    if st.sidebar.button("Logout", width="stretch", key="farmer_market_logout"):
        log_action(farmer_id, "FARMER_MARKET_LOGOUT", "Farmer logged out")
        st.session_state.logged_in = False
        st.session_state.user_id = None
        st.session_state.role = None
        st.rerun()

    st.markdown("## 🌾 AGROW Farmer Market Portal")
    st.caption("Publish your produce, track your listings, review buyer enquiries, check market prices and find agricultural inputs.")

    p1, p2 = st.columns([1, 3])
    with p1:
        if photo_path:
            st.image(photo_path, width=180)
    with p2:
        st.markdown(f"### {farmer_name}")
        st.write(f"**Farmer ID:** {farmer_id}")
        st.write(f"**Location:** {farmer.get('community_village') or '-'}, {farmer.get('lga') or '-'}, {normalize_state_name(farmer.get('state') or '-')}")
        st.write(f"**Primary enterprise:** {farmer.get('primary_crop') or '-'}")
        st.success("✅ Verified agricultural profile")

    listings = fetch_market_listings()
    mine = listings[listings["farmer_id"].astype(str).str.upper() == farmer_id] if not listings.empty else listings
    active_mine = mine[mine["status"] == "AVAILABLE"] if not mine.empty else mine
    enquiries = fetch_market_enquiries()
    my_enquiries = enquiries[enquiries["farmer_id"].astype(str).str.upper() == farmer_id] if not enquiries.empty else enquiries

    m1, m2, m3 = st.columns(3)
    m1.metric("Active Listings", len(active_mine))
    m2.metric("Total Listings", len(mine))
    m3.metric("Buyer Enquiries", len(my_enquiries))

    sell_tab, listings_tab, prices_tab, inputs_tab, enquiries_tab, profile_tab = st.tabs(
        ["➕ Sell Produce", "📦 My Listings", "💰 Market Prices", "🧰 Agricultural Inputs", "💬 Buyer Enquiries", "👤 My Profile"]
    )

    with sell_tab:
        st.markdown("### Create Produce Listing")
        st.caption("Your verified identity and registered location are attached automatically and cannot be changed from this listing form.")
        with st.form("farmer_self_market_listing", clear_on_submit=True):
            c1, c2 = st.columns(2)
            commodity = c1.selectbox("Commodity", MARKET_COMMODITIES, key="farmer_self_commodity")
            product = c2.text_input("Product / Variety", placeholder="e.g. Broiler Chicken, Paddy Rice, Yellow Maize")
            q1, q2, q3 = st.columns(3)
            quantity = q1.number_input("Quantity Available", min_value=0.0, step=1.0)
            unit = q2.selectbox("Unit", MARKET_UNITS, key="farmer_self_unit")
            price = q3.number_input("Asking Price per Unit (₦)", min_value=0.0, step=100.0)
            ready_date = st.date_input("Ready / Harvest Date", key="farmer_self_ready")
            description = st.text_area("Listing Description", placeholder="Quality, grade, average weight, packaging and collection/delivery notes")
            st.info(
                f"📍 Listing location: {farmer.get('community_village') or '-'}, {farmer.get('lga') or '-'}, "
                f"{normalize_state_name(farmer.get('state') or '-')}"
            )
            submit = st.form_submit_button("🚀 Publish to MarketLink", use_container_width=True, type="primary")
            if submit:
                if not product.strip():
                    st.error("Product / variety is required.")
                elif quantity <= 0:
                    st.error("Quantity must be greater than zero.")
                elif price <= 0:
                    st.error("Price must be greater than zero.")
                else:
                    create_market_listing(
                        farmer_id=farmer_id,
                        commodity=commodity,
                        product=product.strip(),
                        quantity=quantity,
                        unit=unit,
                        price=price,
                        ready_date=str(ready_date),
                        state=str(farmer.get("state") or ""),
                        lga=str(farmer.get("lga") or ""),
                        community=str(farmer.get("community_village") or ""),
                        description=description.strip(),
                        created_by=farmer_id,
                    )
                    log_action(farmer_id, "FARMER_SELF_LISTING_CREATED", f"{commodity} | {product.strip()}")
                    st.success("Your produce is now published on AGROW MarketLink.")
                    st.rerun()

    with listings_tab:
        if mine.empty:
            st.info("You have not published any produce yet.")
        else:
            for _, item in mine.iterrows():
                with st.container(border=True):
                    a, b = st.columns([3, 1])
                    a.markdown(f"#### {item.get('product', '-')}")
                    a.write(
                        f"**{item.get('commodity', '-')}** · {float(item.get('available_quantity') or 0):g} {item.get('unit', '')} · "
                        f"**{_money(item.get('price'))} / {item.get('unit', '')}**"
                    )
                    a.caption(f"Status: {item.get('status', '-')} · Ready: {item.get('ready_date') or '-'}")
                    if str(item.get("status", "")) == "AVAILABLE":
                        if b.button("Mark Sold", key=f"farmer_mark_sold_{int(item['id'])}", width="stretch"):
                            update_listing_status(int(item["id"]), "SOLD")
                            log_action(farmer_id, "MARKET_LISTING_SOLD", str(item["id"]))
                            st.rerun()
                        if b.button("Pause", key=f"farmer_pause_{int(item['id'])}", width="stretch"):
                            update_listing_status(int(item["id"]), "PAUSED")
                            st.rerun()
                    elif str(item.get("status", "")) == "PAUSED":
                        if b.button("Re-list", key=f"farmer_relist_{int(item['id'])}", width="stretch"):
                            update_listing_status(int(item["id"]), "AVAILABLE")
                            st.rerun()

    with prices_tab:
        current = fetch_market_listings(status="AVAILABLE")
        if current.empty:
            st.info("The market price board will populate as produce listings are published.")
        else:
            board = (
                current.groupby(["commodity", "product", "state", "unit"], dropna=False)
                .agg(
                    Listings=("id", "count"),
                    Available=("available_quantity", "sum"),
                    Min_Price=("price", "min"),
                    Average_Price=("price", "mean"),
                    Max_Price=("price", "max"),
                )
                .reset_index()
            )
            board["Min Price"] = board["Min_Price"].map(_money)
            board["Average Price"] = board["Average_Price"].map(_money)
            board["Max Price"] = board["Max_Price"].map(_money)
            st.dataframe(
                board[["commodity", "product", "state", "unit", "Listings", "Available", "Min Price", "Average Price", "Max Price"]],
                width="stretch",
                hide_index=True,
            )
            st.caption("MarketLink prices reflect current seller listings and are not an official commodity exchange benchmark.")

    with inputs_tab:
        products = fetch_input_products(status="AVAILABLE")
        if products.empty:
            st.info("No agricultural input products are listed yet.")
        else:
            categories = ["All"] + sorted(products["category"].dropna().unique().tolist())
            category = st.selectbox("Input category", categories, key="farmer_input_category")
            shown = products if category == "All" else products[products["category"] == category]
            for _, item in shown.head(30).iterrows():
                with st.container(border=True):
                    a, b = st.columns([3, 1])
                    a.markdown(f"#### {item.get('product_name', '-')}")
                    a.caption(f"{item.get('category', '-')} · Supplier: {item.get('supplier_name', '-')}")
                    a.write(f"**{_money(item.get('price'))} / {item.get('unit', '')}**")
                    a.write(f"📍 {item.get('lga') or '-'}, {normalize_state_name(item.get('state') or '-')}")
                    phone = str(item.get("supplier_phone") or "")
                    wa_phone = _phone_for_whatsapp(phone)
                    if wa_phone:
                        b.link_button(
                            "💬 Contact Supplier",
                            f"https://wa.me/{wa_phone}?text=Hello%2C%20I%20found%20your%20input%20listing%20on%20AGROW%20MarketLink.",
                            width="stretch",
                        )

    with enquiries_tab:
        if my_enquiries.empty:
            st.info("No buyer enquiries have been recorded for your listings yet.")
        else:
            st.dataframe(
                my_enquiries[["created_at", "product", "buyer_name", "buyer_phone", "quantity_requested", "message", "status"]],
                width="stretch",
                hide_index=True,
            )

    with profile_tab:
        c1, c2 = st.columns([1, 2])
        if photo_path:
            c1.image(photo_path, width=220)
        c2.write(f"**Farmer ID:** {farmer_id}")
        c2.write(f"**Name:** {farmer_name}")
        c2.write(f"**Phone:** {farmer.get('phone_number') or '-'}")
        c2.write(f"**State:** {normalize_state_name(farmer.get('state') or '-')}")
        c2.write(f"**LGA:** {farmer.get('lga') or '-'}")
        c2.write(f"**Crop / Enterprise:** {farmer.get('primary_crop') or '-'}")
        verification_url = (
            f"{get_effective_app_base_url()}?farmer_id={farmer_id}"
            f"&sig={generate_qr_signature(farmer_id)}"
        )
        c2.link_button("✅ Open Public Verified Record", verification_url, width="stretch")
        try:
            farmer_pdf = generate_farmer_id_card_pdf(farmer, photo_path, get_logo_path())
            st.download_button(
                "⬇️ Download Farmer ID Card (PDF)",
                data=farmer_pdf.getvalue(),
                file_name=f"{farmer_name.replace(' ', '_')}_{farmer_id}_Farmer_ID_Card.pdf",
                mime="application/pdf",
                width="stretch",
                type="primary",
                key="farmer_portal_download_id",
            )
        except Exception as exc:
            st.warning(f"Farmer ID card could not be generated at this moment: {exc}")

def show_auth():
    logo_path = get_logo_path()

    st.markdown('<div class="auth-card">', unsafe_allow_html=True)

    if logo_path and os.path.exists(logo_path):
      try:
          col1, col2 = st.columns([1, 6])
          with col1:
            st.image(logo_path, width=90)
      except Exception:
            st.markdown(
                f"""
                <div style="text-align:center; margin-bottom:10px;">
                    <div style="font-size:20px; font-weight:900; color:{PRIMARY_GREEN};">
                        DataDev Limited
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            f"""
            <div style="text-align:center; margin-bottom:10px;">
                <div style="font-size:20px; font-weight:900; color:{PRIMARY_GREEN};">
                    DataDev Limited
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown(
    """
    <div class="top-title">AGROW DIGITAL AGRICULTURE PLATFORM</div>
    <div class="main-title">AGROW — Agricultural Geographic Registration & Operations Workspace</div>
    <div class="sub-title">Farmer Registration, Geographic Intelligence, Market Linkage and Agricultural Operations</div>
    <div class="small-note">Powered by DataDev Limited | Built for Africa’s Agricultural Value Chain</div>
    """,
    unsafe_allow_html=True,
)
    if st.session_state.get("post_signup_redirect", False):
        created_agent_id = st.session_state.get("last_created_agent_id", "")
        st.success(f"Agent account created successfully. Username: {created_agent_id}. Use this username and the password you created to log in.")
        st.session_state["post_signup_redirect"] = False

    st.markdown("### Public Farmer Verification")
    st.caption("Farmers, field agents and input-distribution teams can confirm an enumeration record using a Farmer ID, phone number or the QR code printed on an AGROW Farmer ID card.")
    show_public_farmer_verification(compact=True)
    st.divider()
    market_access_url = f"{get_effective_app_base_url()}?market=1"
    ma1, ma2 = st.columns(2)
    ma1.link_button("🛒 Browse Public MarketLink", market_access_url, width="stretch", type="primary")
    ma2.caption("Consumers, off-takers and processors can browse verified farmer produce without logging in.")
    with st.expander("Preview MarketLink on this page", expanded=False):
        render_public_marketlink()
    st.divider()
    st.markdown("### Secure Workspace Access")

    tab1, tab2, tab3, tab4 = st.tabs(["Staff Login", "Agent Signup", "Password", "🌾 Farmer Market"])

    with tab1:
        default_username = st.session_state.get("prefill_login_username", "")

        login_user = st.text_input(
            "Username",
            value=default_username,
            key="login_username"
        )


        login_pw = st.text_input(
            "Password",
            type="password",
            key="login_password"
        )

        if st.button("🚀 Access Portal", width="stretch"):
            normalized_login_user = str(login_user or "").strip().upper()
            normalized_login_pw = str(login_pw or "").strip()
            user = fetch_user(normalized_login_user)

            if user and str(user["pw"]).strip() == normalized_login_pw:
                st.session_state.logged_in = True
                st.session_state.user_id = normalized_login_user
                st.session_state.role = user["role"]
                log_action(normalized_login_user, "LOGIN", "Successful login")
                st.rerun()
            else:
                st.error("Invalid username or password.")

    with tab2:
        st.markdown("### Agent Registration Details")

        full_name = st.text_input("Full Name", key="signup_full_name")
        phone = st.text_input("Phone Number", max_chars=11, key="signup_phone")
        nin = st.text_input("NIN", max_chars=11, key="signup_nin")
        email = st.text_input("Email Address", key="signup_email")

        state_options = sorted(list(NIGERIA_LGA_MAP.keys()))
        state = st.selectbox("State of Coverage", state_options, key="signup_state")
        lga_options = NIGERIA_LGA_MAP.get(state, [])
        lga_coverage = st.selectbox(
            "LGA of Coverage",
            lga_options if lga_options else ["N/A"],
            key="signup_lga",
        )

        password = st.text_input("Create Password", type="password", key="signup_password")
        confirm_password = st.text_input("Confirm Password", type="password", key="signup_confirm_password")
        invite_code = st.text_input("Organisation Invite Code", key="signup_invite")

        st.markdown("### Agent Photo Capture")
        signup_enable_camera = st.checkbox("Enable Agent Camera", key="signup_enable_camera")
        signup_photo = st.camera_input("Capture Agent Photo", key="signup_photo") if signup_enable_camera else None

        state_prefix = get_state_prefix(state)
        preview_id = generate_agent_id_db(state_prefix)
        st.info(f"Generated Agent Username: {preview_id}")

        if st.button("Create Agent Account", width="stretch"):
            existing = fetch_user(preview_id)

            if not invite_code_matches(invite_code):
                st.error("Invalid organisation invite code. Confirm the local .env value and restart Streamlit.")
            elif existing:
                st.error("Generated agent ID already exists. Please try again.")
            elif not full_name.strip():
                st.error("Full name is required.")
            elif not valid_phone(phone):
                st.error("Phone Number must be exactly 11 digits.")
            elif not valid_nin(nin):
                st.error("NIN must be exactly 11 digits.")
            elif not valid_email(email):
                st.error("A valid email address is required.")
            elif not password.strip():
                st.error("Password is required.")
            elif len(password.strip()) < 6:
                st.error("Password must be at least 6 characters.")
            elif password.strip() != confirm_password.strip():
                st.error("Passwords do not match.")
            elif not signup_enable_camera or signup_photo is None:
                st.error("Agent photo capture is required.")
            else:
                agent_id = preview_id
                clean_password = password.strip()
                insert_user(agent_id, clean_password, "agent", full_name, phone, nin, state, lga_coverage, email)
                saved_user = fetch_user(agent_id)
                if not saved_user or str(saved_user["pw"]).strip() != clean_password:
                    st.error("The account could not be verified after registration. Please try again.")
                    st.stop()
                save_uploaded_photo(signup_photo, str(AGENT_PHOTO_DIR), f"{agent_id}.png")
                log_action(agent_id, "AGENT_CREATED", f"{state} / {lga_coverage}")

                st.session_state["prefill_login_username"] = agent_id
                st.session_state["last_created_agent_id"] = agent_id
                st.session_state["post_signup_redirect"] = True

                signup_keys = [
                    "signup_full_name",
                    "signup_phone",
                    "signup_nin",
                    "signup_email",
                    "signup_state",
                    "signup_lga",
                    "signup_password",
                    "signup_confirm_password",
                    "signup_invite",
                    "signup_enable_camera",
                    "signup_photo",
                ]

                for k in signup_keys:
                    if k in st.session_state:
                        del st.session_state[k]

                st.rerun()

    with tab3:
        st.markdown("### Change Existing Password")

        cp_user = st.text_input("Username", key="cp_user")
        cp_old = st.text_input("Current Password", type="password", key="cp_old")
        cp_new = st.text_input("New Password", type="password", key="cp_new")
        cp_confirm = st.text_input("Confirm New Password", type="password", key="cp_confirm")

        if st.button("Update Password", width="stretch"):
            user = fetch_user(cp_user)
            if not user:
                st.error("User does not exist.")
            elif user["pw"] != cp_old:
                st.error("Current password is incorrect.")
            elif not cp_new.strip():
                st.error("New password cannot be empty.")
            elif len(cp_new) < 6:
                st.error("New password must be at least 6 characters.")
            elif cp_new != cp_confirm:
                st.error("New passwords do not match.")
            else:
                update_user_password(cp_user, cp_new)
                log_action(cp_user, "PASSWORD_CHANGED", "Password updated from login screen")
                st.success("Password updated successfully.")

    with tab4:
        render_farmer_market_login()

    st.markdown("</div>", unsafe_allow_html=True)

def generate_qr_code(url):
    qr = qrcode.make(url)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def show_public_farmer_verification(compact=False, key_prefix="public"):
    if not compact:
        st.markdown("### 🔎 Farmer Verification Portal")
    st.info("Scan an AGROW Farmer ID QR code or search using Farmer ID / Phone Number.")

    query_params = st.query_params
    farmer_id_lookup = query_params.get("farmer_id", "")
    sig = query_params.get("sig", "")
    phone_lookup = ""

    if farmer_id_lookup and sig:
        if not verify_qr_signature(farmer_id_lookup, sig):
            st.error("❌ Invalid or tampered QR code.")
            return

    if not farmer_id_lookup:
        col1, col2 = st.columns(2)

        with col1:
            farmer_id_lookup = st.text_input(
                "Enter Farmer ID",
                placeholder="e.g. AG-20260416074143",
                key=f"{key_prefix}_farmer_id_lookup",
            )

        with col2:
            phone_lookup = st.text_input(
                "Enter Phone Number",
                placeholder="e.g. 08123456789",
                key=f"{key_prefix}_phone_lookup",
            )

        search_col, clear_col = st.columns([4, 1])
        with search_col:
            search_clicked = st.button("🔎 Search Farmer Record", key=f"{key_prefix}_farmer_search", width="stretch", type="primary")
        with clear_col:
            clear_clicked = st.button("Clear", key=f"{key_prefix}_farmer_clear", width="stretch")
        if clear_clicked:
            st.session_state.pop(f"{key_prefix}_farmer_id_lookup", None)
            st.session_state.pop(f"{key_prefix}_phone_lookup", None)
            st.rerun()
        if not search_clicked:
            return

        if not farmer_id_lookup.strip() and not phone_lookup.strip():
            st.warning("Enter Farmer ID or Phone Number.")
            return

    df = fetch_farmers()

    if df.empty:
        st.warning("No farmer records available.")
        return

    record = df.copy()

    if farmer_id_lookup.strip():
        record = record[
            record["farmer_id"].astype(str).str.strip().str.upper()
            == farmer_id_lookup.strip().upper()
        ]
    elif phone_lookup.strip():
        record = record[
            record["phone_number"].astype(str).str.strip()
            == phone_lookup.strip()
        ]

    if record.empty:
        st.error("❌ INVALID FARMER RECORD")
        st.caption(
            "The QR signature is valid, but this Farmer ID is not present in the active database. "
            "This commonly occurs when a farmer was registered in an older local/temporary database before PostgreSQL persistence was enabled."
        )
        return

    farmer = record.iloc[0]
    farmer_id_value = str(farmer.get("farmer_id", "") or "").strip()
    photo_path = resolve_farmer_photo_path(
        farmer.get("photo_path", ""),
        farmer_id_value,
    )

    st.success("✅ Farmer Verified")

    col1, col2 = st.columns([1, 2])

    with col1:
        if photo_path:
            st.image(photo_path, width=220)
        else:
            st.warning("No photo available on this server for this farmer record.")

    with col2:
        st.markdown(f"""
**Farmer ID:** {farmer.get('farmer_id', '-')}  
**Name:** {farmer.get('farmer_full_name', '-')}  
**Phone:** {farmer.get('phone_number', '-')}  
**State:** {normalize_state_name(farmer.get('state', '-'))}  
**LGA:** {farmer.get('lga', '-')}  
**Crop:** {farmer.get('primary_crop', '-')}  
**NIN Status:** {farmer.get('nin_status', '-')}  
        """)

    pdf_file = generate_farmer_id_card_pdf(
        selected_row=farmer,
        photo_path=photo_path,
        logo_path=get_logo_path(),
    )

    st.download_button(
        label="⬇️ Download Farmer ID Card",
        data=pdf_file.getvalue(),
        file_name=f"{safe_filename(farmer.get('farmer_full_name', 'Farmer'))}_{safe_filename(farmer.get('farmer_id', 'farmer'))}_ID_Card.pdf",
        mime="application/pdf",
        width="stretch",
    )

# =========================================================
# 9. APP BODY
# =========================================================
public_market_mode = str(st.query_params.get("market", "")).strip().lower() in {"1", "true", "yes"}
public_listing_id = st.query_params.get("listing_id", "")

if not st.session_state.logged_in and public_market_mode:
    st.markdown(
        """
        <div style="text-align:center; margin-bottom:16px;">
            <div style="font-size:clamp(24px,4vw,34px);font-weight:900;color:#004B87;">AGROW MarketLink</div>
            <div style="font-size:clamp(13px,2vw,17px);font-weight:700;color:#1B5E20;">Verified agricultural produce, inputs and market-price intelligence</div>
        </div>
        """, unsafe_allow_html=True
    )
    if public_listing_id:
        render_public_listing_detail(public_listing_id)
        st.link_button("← Browse all MarketLink listings", f"{get_effective_app_base_url()}?market=1", width="stretch")
    else:
        render_public_marketlink()
    st.divider()
    st.link_button("🔎 Farmer Verification / Secure Access", get_effective_app_base_url(), width="stretch")
    st.stop()

if not st.session_state.logged_in:
    show_auth()

else:
    user_id = st.session_state.user_id
    role = st.session_state.role

    if role == "farmer":
        render_farmer_market_portal(user_id)
        st.stop()

    user_meta = fetch_user(user_id) or {}

    # continue dashboard here...

    # logged-in dashboard continues here

    if not user_id or not role:
     st.session_state.logged_in = False
     st.session_state.user_id = None
     st.session_state.role = None
     st.warning("Session expired. Please log in again.")
     show_auth()
     st.stop()

    def is_db_available():
        try:
            conn = get_connection()
            conn.close()
            return True
        except Exception:
            return False

    logo_path = get_logo_path()

    if logo_path and os.path.exists(logo_path):
        try:
            col1, col2, col3 = st.sidebar.columns([1, 1.3, 1])
            with col2:
                st.image(logo_path, width=95)

            st.sidebar.markdown(
                f"""
                <div style="text-align:center; margin-top:-2px; margin-bottom:14px;">
                    <div style="font-size:17px; font-weight:800; color:{PRIMARY_GREEN};">
                        DataDev Limited
                    </div>
                    <div style="font-size:10px; color:#666;">
                        Digital Systems for Development
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        except Exception:
            st.sidebar.markdown(
                f"""
                <div style="text-align:center; margin-top:10px; margin-bottom:12px;">
                    <div style="font-size:20px; font-weight:900; color:{PRIMARY_GREEN};">
                        DataDev Limited
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
    else:
        st.sidebar.markdown(
            f"""
            <div style="text-align:center; margin-top:10px; margin-bottom:12px;">
                <div style="font-size:20px; font-weight:900; color:{PRIMARY_GREEN};">
                    DataDev Limited
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.sidebar.subheader(f"👤 User: {user_id or 'Guest'}")
    st.sidebar.write(f"**Role:** {(role or 'Guest').title()}")

    if role == "agent":
        st.sidebar.write(f"**Coverage State:** {normalize_state_name(user_meta.get('state', '-'))}")
        st.sidebar.write(f"**Coverage LGA:** {user_meta.get('lga_coverage', '-')}")

    if st.sidebar.button("Logout", width="stretch"):
        log_action(user_id, "LOGOUT", "User logged out")
        st.session_state.logged_in = False
        st.session_state.user_id = None
        st.session_state.role = None
        st.rerun()

    with st.sidebar.expander("🔑 Change Password"):
        old_pw = st.text_input("Current Password", type="password", key="side_old_pw")
        new_pw = st.text_input("New Password", type="password", key="side_new_pw")
        confirm_pw = st.text_input("Confirm New Password", type="password", key="side_confirm_pw")

        if st.button("Save New Password", width="stretch"):
            user = fetch_user(user_id)
            if not user or user["pw"] != old_pw:
                st.error("Current password is incorrect.")
            elif not new_pw.strip():
                st.error("New password cannot be empty.")
            elif len(new_pw) < 6:
                st.error("New password must be at least 6 characters.")
            elif new_pw != confirm_pw:
                st.error("New passwords do not match.")
            else:
                update_user_password(user_id, new_pw)
                log_action(user_id, "PASSWORD_CHANGED", "Password updated from sidebar")
                st.success("Password changed successfully.")

    state_options = sorted(list(NIGERIA_LGA_MAP.keys()))

    if role == "admin":
        selected_state = st.sidebar.selectbox("Region Focus", ["All Nigeria"] + state_options)
    else:
        selected_state = user_meta.get("state", "All Nigeria")
        st.sidebar.info(f"Access restricted to {normalize_state_name(selected_state)}")

    st.markdown(
        """
        <div style="text-align:center; margin-bottom:18px; padding: 0 10px;">
            <div style="font-size:clamp(22px, 4vw, 32px); font-weight:900; color:#004B87; line-height:1.2;">
                AGROW — Agricultural Geographic Registration & Operations Workspace
            </div>
            <div style="font-size:clamp(13px, 2.2vw, 16px); font-weight:700; color:#1B5E20; margin-top:6px; line-height:1.4;">
                A unified workspace for farmer identity, geospatial registration, field operations, market access and agricultural intelligence
            </div>
            <div style="font-size:clamp(11px, 1.8vw, 14px); color:#555; margin-top:6px; line-height:1.5;">
                Connecting verified farmer records, geographic field intelligence, input distribution and accountable agricultural operations
            </div>
            <div style="font-size:clamp(10px, 1.6vw, 13px); color:#6A6A6A; margin-top:5px; line-height:1.45;">
                Built for governments, NGOs, development programmes, cooperatives, agribusinesses, research organisations and other agricultural institutions
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.info("⚠️ Demo Mode: Sample data displayed for presentation purposes")
    st.caption("Powered by DataDev Limited | Agricultural Value-Chain Infrastructure")

    if is_db_available():
        st.success("🟢 ONLINE MODE: Data syncing to central database")
    else:
        st.warning("🟡 OFFLINE MODE: Data stored locally and will sync later")

    if role == "admin":
        st.markdown("### 🌐 AGROW Platform Access")

        qr_img = generate_qr_code(APP_BASE_URL)

        qr_col1, qr_col2 = st.columns([1, 4])

        with qr_col1:
            st.image(qr_img, width=150)

        with qr_col2:
            st.caption("Scan this platform QR to open the public verification and secure workspace landing page on another device.")
            st.download_button(
                label="⬇️ Download QR Code",
                data=qr_img.getvalue(),
                file_name="agrow_platform_qr.png",
                mime="image/png",
                width="content",
            )

        st.divider()

    st.caption(f"Last sync check: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # -------------------------
    # Dashboard data
    # -------------------------
    df = fetch_farmers()

    if selected_state != "All Nigeria" and not df.empty:
        df = df[df["state"] == selected_state]

    if role == "agent" and not df.empty:
        df = df[df["agent_id"] == user_id]

    total_beneficiaries = len(df)
    total_verified = len(df[df["nin_status"] == "Verified"]) if not df.empty else 0
    total_land = float(df["farm_size_ha"].sum()) if not df.empty else 0.0
    verification_rate = round((total_verified / total_beneficiaries) * 100, 1) if total_beneficiaries > 0 else 0.0

    # =========================================================
    # TAB LAYOUT
    # =========================================================
    tab_dashboard, tab_farmer_registration, tab_distribution, tab_verification, tab_marketlink, tab_analytics = st.tabs(
        ["📊 Dashboard", "📝 Farmer Registration", "📦 Distribution", "🔎 Farmer Verification", "🛒 MarketLink", "📈 Analytics"]
    )

    # =========================================================
    # TAB 1: DASHBOARD
    # =========================================================
    with tab_dashboard:
        total_agents = len(fetch_all_agents()) if role == "admin" else 0

        refresh_col, refresh_note_col = st.columns([1, 3])
        with refresh_col:
            if st.button("↻ Refresh Live Data", key="dashboard_refresh_live_data", width="stretch"):
                st.rerun()
        with refresh_note_col:
            st.caption("Headline registration metrics and sync activity refresh automatically every 10 seconds. Use Refresh Live Data for an immediate full dashboard update.")

        render_live_dashboard_overview(selected_state, role, user_id)

        if role == "admin":
            st.markdown("### Admin Monitoring Snapshot")
            a1, a2 = st.columns(2)
            a3, a4 = st.columns(2)

            a1.info(f"Registered Agents: {total_agents}")
            a2.info("Target Registered Farmers: 1,000,000")
            a3.info("Target States Covered: 36 + FCT")
            a4.info("Target Verification Rate: 95%")
        else:
            st.markdown("### Field Agent Monitoring Snapshot")
            st.info(f"Coverage State: {normalize_state_name(user_meta.get('state', '-'))}")
            st.info(f"Coverage LGA: {user_meta.get('lga_coverage', '-')}")
            st.info("Target Agent Uptime: 99%")

        # ---------------------------------------------------------
        # Dashboard reporting access
        # ---------------------------------------------------------
        st.markdown("### Data & Reporting")
        st.caption(
            "Download the complete non-image farmer register available within your current access scope. "
            "The export retains identity, contact, location, crop, farm-size, input-distribution, verification, "
            "GPS and enumerator fields for programme monitoring and further analysis."
        )

        dashboard_export_df = fetch_complete_farmer_registry()
        if not dashboard_export_df.empty:
            # Respect the same state/agent scope already applied to the signed-in dashboard.
            dashboard_fid_col = next(
                (c for c in dashboard_export_df.columns if _canonical_farmer_column_name(c) == "Farmer_ID"),
                None,
            )
            if dashboard_fid_col and not df.empty and "farmer_id" in df.columns:
                dashboard_wanted_ids = set(df["farmer_id"].fillna("").astype(str))
                dashboard_export_df = dashboard_export_df[
                    dashboard_export_df[dashboard_fid_col].fillna("").astype(str).isin(dashboard_wanted_ids)
                ].copy()
            elif df.empty:
                dashboard_export_df = dashboard_export_df.iloc[0:0].copy()

        if dashboard_export_df.empty:
            st.info("No farmer records are currently available within this access scope for export.")
        else:
            report_col1, report_col2 = st.columns(2)
            with report_col1:
                table_to_excel_download(dashboard_export_df, "agrow_master_registry.xlsx", key="dashboard_master_excel")
            with report_col2:
                table_to_csv_download(dashboard_export_df, "agrow_master_registry.csv", key="dashboard_master_csv")
            st.caption(f"Records available in this export: {len(clean_registry_export(dashboard_export_df)):,}")

        st.divider()

        st.subheader("Geospatial Mapping")
        if not df.empty:
            map_df = df.rename(columns={"latitude": "lat", "longitude": "lon"})
            st.map(map_df)
        else:
            st.info("No field data available for mapping.")

        st.divider()

        c1, c2 = st.columns(2)

        with c1:
            st.subheader("Registrations by State")
            if not df.empty:
                state_counts = df["state"].value_counts().reset_index()
                state_counts.columns = ["State", "Count"]
                state_counts["State"] = state_counts["State"].apply(normalize_state_name)
                st.plotly_chart(px.bar(state_counts, x="State", y="Count"), width="stretch")
            else:
                st.info("No state data")

        with c2:
            st.subheader("Registrations by Agent")
            if not df.empty:
                agent_counts = df["agent_id"].value_counts().reset_index()
                agent_counts.columns = ["Agent_ID", "Count"]
                st.plotly_chart(px.bar(agent_counts, x="Agent_ID", y="Count"), width="stretch")
            else:
                st.info("No agent data")

        st.divider()

        d1, d2 = st.columns(2)

        with d1:
            st.subheader("NIN Status Distribution")
            if not df.empty:
                nin_counts = df["nin_status"].value_counts().reset_index()
                nin_counts.columns = ["NIN_Status", "Count"]
                st.plotly_chart(
                    px.pie(nin_counts, names="NIN_Status", values="Count", hole=0.45),
                    width="stretch",
                )
            else:
                st.info("No NIN data")

        with d2:
            st.subheader("Photo Capture Status")
            if not df.empty:
                photo_counts = df["photo_status"].value_counts().reset_index()
                photo_counts.columns = ["Photo_Status", "Count"]
                st.plotly_chart(
                    px.pie(photo_counts, names="Photo_Status", values="Count", hole=0.45),
                    width="stretch",
                )
            else:
                st.info("No photo data")

        st.divider()

        st.subheader("Input Distribution Summary")
        if not df.empty:
            input_counts = df["input_distributed"].fillna("").str.split(", ").explode()
            input_counts = input_counts[input_counts != ""].value_counts().reset_index()
            input_counts.columns = ["Input_Type", "Count"]
            st.plotly_chart(px.bar(input_counts, x="Input_Type", y="Count"), width="stretch")
        else:
            st.info("No input data")

        st.divider()

        st.subheader("Recent Farmer Registrations")
        if not df.empty:
            recent_df = clean_registry_export(df)
            recent_screen_columns = [
                "Farmer_ID", "Registration_Date", "Agent_ID", "Farmer_Full_Name",
                "Gender", "Date_of_Birth", "Phone_Number", "State", "LGA",
                "Primary_Crop", "NIN_Status", "Photo_Captured",
            ]
            recent_screen_columns = [
                c for c in recent_screen_columns if c in recent_df.columns
            ]
            st.dataframe(
                recent_df.head(10)[recent_screen_columns],
                width="stretch",
                hide_index=True,
            )
        else:
            st.info("No farmer records yet")

        st.divider()

        render_live_sync_monitor(selected_state, role, user_id)

    # =========================================================
    # TAB 2: REGISTER FARMER
    # =========================================================
    @st.fragment
    def render_farmer_registration():
        st.info("Complete each section carefully and click 'Sync Secure Record' only after all required information has been entered.")

        if "registration_success" in st.session_state:
            st.success(st.session_state["registration_success"])
            del st.session_state["registration_success"]

        form_v = st.session_state.get("farmer_form_version", 0)

        input_options = [
            "Rice Seeds",
            "Maize Seeds",
            "Cassava Stems",
            "Sorghum Seeds",
            "Soybean Seeds",
            "Groundnut Seeds",
            "NPK",
            "Urea",
            "Organic Fertilizer",
            "Herbicide",
            "Pesticide",
            "Insecticide",
            "Sprayer",
            "Water Pump",
            "Extension Support",
        ]

        # Every successful submission increments this version. Streamlit then
        # creates a fresh set of widget keys, giving the agent a clean form.
        form_v = st.session_state.get("farmer_form_version", 0)

        st.markdown("### Section 1: Farmer Identity")
        farmer_full_name = st.text_input("Farmer Full Name", key=f"farmer_full_name_{form_v}")
        identity_col1, identity_col2 = st.columns(2)
        with identity_col1:
            gender = st.selectbox("Gender", ["Male", "Female"], key=f"gender_{form_v}")
        with identity_col2:
            dob = st.date_input("Date of Birth", key=f"dob_{form_v}")

        st.markdown("---")
        st.markdown("### Section 2: Contact Information")
        contact_col1, contact_col2 = st.columns(2)
        with contact_col1:
            phone_number = st.text_input("Phone Number", max_chars=11, key=f"phone_number_{form_v}")
        with contact_col2:
            alternate_phone = st.text_input(
                "Alternate Phone Number",
                max_chars=11,
                placeholder="Optional 11-digit phone number",
                key=f"alternate_phone_{form_v}",
            )
        email_address = st.text_input("Email Address", key=f"email_address_{form_v}")

        st.markdown("---")
        st.markdown("### Section 3: Coverage and Location")
        if role == "admin":
            farmer_state = st.selectbox("State", state_options, key=f"farmer_state_{form_v}")
        else:
            farmer_state = user_meta.get("state", "")
            st.text_input("State", value=farmer_state, disabled=True, key=f"farmer_state_display_{form_v}")

        lga_list = NIGERIA_LGA_MAP.get(farmer_state, [])
        if role == "admin":
            farmer_lga = st.selectbox("LGA", lga_list if lga_list else ["N/A"], key=f"farmer_lga_{form_v}")
        else:
            assigned_lga = user_meta.get("lga_coverage", "")
            farmer_lga = assigned_lga if assigned_lga in lga_list or assigned_lga else (lga_list[0] if lga_list else "N/A")
            st.text_input("LGA", value=farmer_lga, disabled=True, key=f"farmer_lga_display_{form_v}")

        location_col1, location_col2 = st.columns(2)
        with location_col1:
            ward = st.text_input("Ward", key=f"ward_{form_v}")
        with location_col2:
            community = st.text_input("Community / Village", key=f"community_{form_v}")
        residential_address = st.text_area("Residential Address", key=f"residential_address_{form_v}")

        st.markdown("---")
        st.markdown("### Section 4: Farm Profile")
        crop_col1, crop_col2 = st.columns(2)
        with crop_col1:
            primary_crop = st.selectbox(
                "Primary Crop",
                ["Rice", "Maize", "Cassava", "Sorghum", "Soybean", "Groundnut", "Millet"],
                key=f"primary_crop_{form_v}",
            )
        with crop_col2:
            secondary_crop = st.selectbox(
                "Secondary Crop",
                ["None", "Rice", "Maize", "Cassava", "Sorghum", "Soybean", "Groundnut", "Millet"],
                key=f"secondary_crop_{form_v}",
            )
        farm_size = st.number_input("Farm Size (Hectares)", min_value=0.0, step=0.1, key=f"farm_size_{form_v}")

        st.markdown("---")
        st.markdown("### Section 5: Support Delivered")
        input_list = st.multiselect("Inputs Distributed", input_options, key=f"input_list_{form_v}")
        quantity_units = len(input_list)
        st.caption(f"Total input types selected: {quantity_units}")

        st.markdown("---")
        st.markdown("### Section 6: NIN Verification")
        nin_col1, nin_col2 = st.columns(2)
        with nin_col1:
            nin = st.text_input("NIN", max_chars=11, key=f"nin_{form_v}")
        with nin_col2:
            nin_status = st.selectbox(
                "NIN Verification Status",
                ["Verified", "Pending", "Rejected"],
                key=f"nin_status_{form_v}",
            )

        st.markdown("---")
        st.markdown("### Section 7: Identification Document")
        id_col1, id_col2 = st.columns(2)
        with id_col1:
            id_type = st.selectbox(
                "ID Type",
                ["NIN Slip", "Voter Card", "Driver's License", "National ID Card", "Other"],
                key=f"id_type_{form_v}",
            )
        with id_col2:
            id_number = st.text_input("ID Number", max_chars=20, key=f"id_number_{form_v}")

        st.markdown("---")
        st.markdown("### Section 8: Automatic Farm Geo-tagging")
        st.caption(
            "Capture the device GPS position at the farm. The browser will request location permission the first time. "
            "Coordinates remain editable when a field agent needs to correct a weak GPS reading."
        )

        location = streamlit_geolocation()
        lat_key = f"latitude_{form_v}"
        lon_key = f"longitude_{form_v}"
        accuracy_key = f"gps_accuracy_{form_v}"

        if isinstance(location, dict):
            detected_lat = location.get("latitude")
            detected_lon = location.get("longitude")
            detected_accuracy = location.get("accuracy")
            if detected_lat is not None and detected_lon is not None:
                st.session_state[lat_key] = float(detected_lat)
                st.session_state[lon_key] = float(detected_lon)
                st.session_state[accuracy_key] = detected_accuracy
                accuracy_text = f" ±{float(detected_accuracy):.1f} m" if detected_accuracy is not None else ""
                st.success(f"GPS captured: {float(detected_lat):.6f}, {float(detected_lon):.6f}{accuracy_text}")
        else:
            st.info("Tap the location button and allow GPS access before submitting this farmer record.")

        geo_col1, geo_col2 = st.columns(2)
        with geo_col1:
            latitude = st.number_input("Latitude", format="%.6f", key=lat_key)
        with geo_col2:
            longitude = st.number_input("Longitude", format="%.6f", key=lon_key)

        if role == "agent" and latitude and longitude:
            geofence_result = gps_matches_assigned_state(farmer_state, float(latitude), float(longitude))
            if geofence_result is False:
                st.error(f"GPS is outside the broad {normalize_state_name(farmer_state)} operating envelope.")
            elif geofence_result is True:
                st.success(f"GPS is consistent with the assigned state: {normalize_state_name(farmer_state)}.")
            else:
                st.info("State and LGA are locked to the agent assignment. Precise boundary validation requires official GIS polygons.")

        st.markdown("---")
        st.markdown("### Section 9: Farmer Photo Capture")
        enable_camera_main = st.checkbox("Enable Camera", key=f"enable_camera_main_{form_v}")
        photo_capture = None
        if enable_camera_main:
            photo_capture = st.camera_input("Capture Farmer Photo", key=f"photo_capture_{form_v}")

        st.markdown("---")
        st.markdown("### Section 10: Enumerator Remarks and Confirmation")
        remarks = st.text_area("Enumerator Remarks", key=f"remarks_{form_v}")
        record_confirmed = st.checkbox(
            "I confirm that the information above has been reviewed with the farmer and is ready for submission.",
            key=f"record_confirmed_{form_v}",
        )

        st.markdown("---")
        st.markdown("### Section 11: Submission")
        submitted = st.button("Sync Secure Record", width="stretch", key=f"submit_farmer_{form_v}")

        if submitted:
            if not farmer_full_name.strip():
                st.error("Farmer full name is required.")
            elif not valid_phone(phone_number):
                st.error("Phone Number must be exactly 11 digits.")
            elif alternate_phone and not valid_phone(alternate_phone):
                st.error("Alternate Phone Number must be exactly 11 digits.")
            elif not valid_nin(nin):
                st.error("NIN must be exactly 11 digits.")
            elif not valid_id_number(id_number):
                st.error("ID Number must be between 1 and 20 characters.")
            elif not farmer_lga or farmer_lga == "N/A":
                st.error("LGA is required.")
            elif len(input_list) == 0:
                st.error("Please select at least one distributed input.")
            elif photo_capture is None:
                st.error("Farmer photo capture is required.")
            elif latitude == 0.0 and longitude == 0.0:
                st.error("Capture the farm GPS location before submission. Allow browser location access and try again.")
            elif role == "agent" and gps_matches_assigned_state(farmer_state, float(latitude), float(longitude)) is False:
                st.error("Submission blocked: GPS is outside the agent's assigned-state operating envelope.")
            elif not record_confirmed:
                st.error("Please confirm that the farmer's information has been reviewed before submission.")
            elif farmer_exists_today(phone_number, nin, farmer_full_name):
                st.warning("This farmer appears to have already been registered today.")
            else:
                now = datetime.now()
                farmer_id = f"AG-{now.strftime('%Y%m%d%H%M%S')}"

                photo_path = ""
                if photo_capture is not None:
                    photo_filename = f"{farmer_id}.jpg"
                    photo_path = save_uploaded_photo(
                        photo_capture,
                        folder=FARMER_PHOTO_DIR,
                        filename=photo_filename,
                    )
                    photo_path = resolve_farmer_photo_path(photo_path, farmer_id)
                    if not photo_path:
                        st.error("Farmer photo could not be saved. The farmer record has not been submitted.")
                        st.stop()

                row = {
                    "Farmer_ID": farmer_id,
                    "Registration_Date": now.strftime("%Y-%m-%d %H:%M:%S"),
                    "Agent_ID": user_id,
                    "Farmer_Full_Name": farmer_full_name,
                    "Gender": gender,
                    "Date_of_Birth": str(dob),
                    "Phone_Number": phone_number,
                    "Alternate_Phone": alternate_phone,
                    "Email_Address": email_address,
                    "NIN": nin,
                    "State": farmer_state,
                    "LGA": farmer_lga,
                    "Ward": ward,
                    "Community_Village": community,
                    "Residential_Address": residential_address,
                    "Primary_Crop": primary_crop,
                    "Secondary_Crop": secondary_crop,
                    "Farm_Size_Ha": farm_size,
                    "Input_Distributed": ", ".join(input_list),
                    "Quantity_Units": quantity_units,
                    "NIN_Status": nin_status,
                    "ID_Type": id_type,
                    "ID_Number": id_number,
                    "Latitude": latitude,
                    "Longitude": longitude,
                    "Enumerator_Remarks": remarks,
                    "Photo_Path": photo_path,
                    "Photo_Status": "Captured" if photo_path else "No Photo",
                }

                if is_db_available():
                    insert_farmer(row)
                    log_action(user_id, "FARMER_REGISTERED", farmer_id)
                    success_text = "✅ Farmer record saved directly to the main database."
                else:
                    save_to_offline_queue(row)
                    log_action(user_id, "FARMER_QUEUED", farmer_id)
                    success_text = "✅ Offline Mode: farmer record saved locally and queued for sync."

                # Preserve the confirmation across the rerun, while assigning a
                # new widget-key version so the registration screen is blank.
                st.session_state["registration_success"] = f"{success_text} Farmer ID: {farmer_id}"
                st.session_state["farmer_form_version"] = form_v + 1
                st.rerun(scope="app")


    with tab_farmer_registration:
        render_farmer_registration()

    # =========================================================
    # TAB 3: DISTRIBUTION
    # =========================================================
    with tab_distribution:
        st.subheader("Input Breakdown")

        if not df.empty:
            input_counts = df["input_distributed"].fillna("").str.split(", ").explode()
            input_counts = input_counts[input_counts != ""].value_counts().reset_index()
            input_counts.columns = ["Input_Type", "Count"]

            pie_fig = px.pie(
                input_counts,
                names="Input_Type",
                values="Count",
                hole=0.5,
            )
            # Soft agricultural green canvas gives the donut definition without
            # competing with Plotly's categorical segment colours.
            pie_fig.update_layout(
                paper_bgcolor="#F1F7F2",
                plot_bgcolor="#F1F7F2",
                margin=dict(l=28, r=28, t=32, b=28),
                legend=dict(
                    bgcolor="rgba(255,255,255,0.72)",
                    bordercolor="#D5E4D8",
                    borderwidth=1,
                ),
                font=dict(color="#26352B"),
            )
            pie_fig.update_traces(
                marker=dict(line=dict(color="#FFFFFF", width=1.5)),
                textfont=dict(color="#26352B"),
            )
            st.plotly_chart(pie_fig, width="stretch")
        else:
            st.info("No input distribution data available.")

        st.subheader("📋 Master Registry Database")

        if st.button("Reset Filters", key="reset_distribution_filters"):
            st.session_state["search_name"] = ""
            st.session_state["search_agent"] = ""
            st.session_state["filter_nin_status"] = "All"
            st.session_state["filter_photo_status"] = "All"
            st.session_state["filter_state"] = "All"
            st.session_state["filter_lga"] = "All"
            st.session_state["search_phone"] = ""

            if "photo_preview_select" in st.session_state:
                del st.session_state["photo_preview_select"]

            st.rerun()

        search_col1, search_col2, search_col3, search_col4 = st.columns(4)

        with search_col1:
            search_name = st.text_input("Search Farmer Name", key="search_name")

        with search_col2:
            search_agent = st.text_input("Search Agent ID", key="search_agent", placeholder="e.g. FCT-01")

        with search_col3:
            filter_nin_status = st.selectbox(
                "Filter NIN Status",
                ["All", "Verified", "Pending", "Rejected"],
                key="filter_nin_status"
            )

        with search_col4:
            filter_photo_status = st.selectbox(
                "Filter Photo Status",
                ["All", "Captured", "No Photo"],
                key="filter_photo_status"
            )

        filter_col1, filter_col2, filter_col3 = st.columns(3)

        with filter_col1:
            state_filter_options = ["All"] + sorted(df["state"].dropna().unique().tolist()) if not df.empty else ["All"]
            filter_state = st.selectbox("Filter State", state_filter_options, key="filter_state")

        with filter_col2:
            lga_filter_options = ["All"]
            if not df.empty and filter_state != "All":
                lga_filter_options += sorted(df[df["state"] == filter_state]["lga"].dropna().unique().tolist())
            elif not df.empty:
                lga_filter_options += sorted(df["lga"].dropna().unique().tolist())
            filter_lga = st.selectbox("Filter LGA", lga_filter_options, key="filter_lga")

        with filter_col3:
            search_phone = st.text_input("Search Phone Number", key="search_phone")

        filtered_df = df.copy()

        if search_name:
            filtered_df = filtered_df[
                filtered_df["farmer_full_name"].astype(str).str.contains(search_name, case=False, na=False)
            ]

        if search_agent:
            filtered_df = filtered_df[
                filtered_df["agent_id"].astype(str).str.contains(search_agent, case=False, na=False)
            ]

        if search_phone:
            filtered_df = filtered_df[
                filtered_df["phone_number"].astype(str).str.contains(search_phone, case=False, na=False)
            ]

        if filter_nin_status != "All":
            filtered_df = filtered_df[filtered_df["nin_status"] == filter_nin_status]

        if filter_photo_status != "All":
            filtered_df = filtered_df[filtered_df["photo_status"] == filter_photo_status]

        if filter_state != "All":
            filtered_df = filtered_df[filtered_df["state"] == filter_state]

        if filter_lga != "All":
            filtered_df = filtered_df[filtered_df["lga"] == filter_lga]

        st.caption(f"Showing {len(filtered_df)} record(s)")

        if filtered_df.empty:
            st.warning("No farmer record matched the current search/filter selection. Click 'Reset Filters' and try again.")

        display_df = filtered_df.rename(columns={
            "farmer_id": "Farmer_ID",
            "registration_date": "Registration_Date",
            "agent_id": "Agent_ID",
            "farmer_full_name": "Farmer_Full_Name",
            "gender": "Gender",
            "date_of_birth": "Date_of_Birth",
            "phone_number": "Phone_Number",
            "alternate_phone": "Alternate_Phone",
            "email_address": "Email_Address",
            "nin": "NIN",
            "state": "State",
            "lga": "LGA",
            "ward": "Ward",
            "community_village": "Community_Village",
            "residential_address": "Residential_Address",
            "primary_crop": "Primary_Crop",
            "secondary_crop": "Secondary_Crop",
            "farm_size_ha": "Farm_Size_Ha",
            "input_distributed": "Input_Distributed",
            "quantity_units": "Quantity_Units",
            "nin_status": "NIN_Status",
            "id_type": "ID_Type",
            "id_number": "ID_Number",
            "latitude": "Latitude",
            "longitude": "Longitude",
            "enumerator_remarks": "Enumerator_Remarks",
            "photo_status": "Photo_Status",
            "photo_path": "Photo_Path",
        })

        preferred_columns = [
            "Farmer_ID", "Registration_Date", "Agent_ID", "Farmer_Full_Name", "Gender",
            "Date_of_Birth", "Phone_Number", "Alternate_Phone", "Email_Address", "NIN",
            "State", "LGA", "Ward", "Community_Village", "Residential_Address",
            "Primary_Crop", "Secondary_Crop", "Farm_Size_Ha", "Input_Distributed",
            "Quantity_Units", "NIN_Status", "ID_Type", "ID_Number", "Latitude",
            "Longitude", "Enumerator_Remarks", "Photo_Status", "Photo_Path",
        ]

        # Keep every database registration field for CSV/XLSX export.
        # The screen remains compact below; clean_registry_export removes only image/binary fields.

        screen_columns = [
            "Farmer_ID", "Registration_Date", "Agent_ID", "Farmer_Full_Name", "Gender",
            "Phone_Number", "State", "LGA", "Primary_Crop", "Quantity_Units",
            "NIN_Status", "Photo_Status",
        ]

        available_screen_columns = [col for col in screen_columns if col in display_df.columns]
        st.dataframe(display_df[available_screen_columns], width="stretch")

        st.markdown("### Farmer ID Preview")

        if "photo_path" in filtered_df.columns:
            photo_records = filtered_df[
                filtered_df["photo_path"].notna() & (filtered_df["photo_path"] != "")
            ]
        else:
            photo_records = pd.DataFrame()

        if not photo_records.empty:
            photo_options = (
                photo_records["farmer_full_name"].fillna("Unknown").astype(str)
                + " | "
                + photo_records["farmer_id"].fillna("").astype(str)
            ).tolist()

            selected_photo_label = st.selectbox(
                "Select farmer photo to preview",
                photo_options,
                key="photo_preview_select"
            )

            selected_index = photo_options.index(selected_photo_label)
            selected_row = photo_records.iloc[selected_index]
            selected_photo_path = resolve_farmer_photo_path(
                selected_row.get("photo_path", ""),
                selected_row.get("farmer_id", ""),
            )

            if selected_photo_path:
                col1, col2 = st.columns([1, 2])

                with col1:
                    st.image(selected_photo_path, width=220)

                with col2:
                    st.markdown(f"""
**Farmer ID:** {selected_row['farmer_id']}  
**Name:** {selected_row['farmer_full_name']}  
**Phone:** {selected_row['phone_number']}  
**State:** {normalize_state_name(selected_row['state'])}  
**LGA:** {selected_row['lga']}  
**Crop:** {selected_row.get('primary_crop', '-')}  
**Photo Status:** {selected_row['photo_status']}
                    """)

                farmer_id_value = str(selected_row['farmer_id'])
                verification_url = f"{get_effective_app_base_url()}?farmer_id={farmer_id_value}&sig={generate_qr_signature(farmer_id_value)}"
                farmer_qr = generate_qr_code(verification_url)
                farmer_qr_identity = build_farmer_qr_identity_image(selected_row, verification_url)
                qr1, qr2 = st.columns([1, 3])
                with qr1:
                    st.image(farmer_qr, width=150, caption="Individual Farmer QR")
                with qr2:
                    st.caption("Scan during input distribution or field verification to open this farmer's authenticated enumeration record.")
                    farmer_id_pdf = generate_farmer_id_card_pdf(
                        selected_row=selected_row,
                        photo_path=selected_photo_path,
                        logo_path=get_logo_path(),
                    )
                    st.download_button(
                        "Download Farmer ID Card (PDF)",
                        data=farmer_id_pdf.getvalue(),
                        file_name=f"{safe_filename(selected_row.get('farmer_full_name', 'Farmer'))}_{safe_filename(farmer_id_value)}_Farmer_ID_Card.pdf",
                        mime="application/pdf",
                        key=f"download_farmer_id_pdf_{farmer_id_value}",
                        width="content",
                    )
                    st.caption(f"QR verification address: {verification_url}")
            else:
                st.warning("Photo file path exists in record, but image file was not found on disk.")
        else:
            st.info("No captured farmer photos available for preview yet.")

        # Full analytical export comes directly from the authoritative farmers table,
        # not from the compact screen projection.  Preserve the current filter by Farmer_ID.
        full_export_df = fetch_complete_farmer_registry()
        if not full_export_df.empty and not filtered_df.empty:
            fid_col = next((c for c in full_export_df.columns if _canonical_farmer_column_name(c) == "Farmer_ID"), None)
            if fid_col and "farmer_id" in filtered_df.columns:
                wanted_ids = set(filtered_df["farmer_id"].fillna("").astype(str))
                full_export_df = full_export_df[full_export_df[fid_col].fillna("").astype(str).isin(wanted_ids)].copy()
        elif full_export_df.empty:
            full_export_df = display_df

        download_col1, download_col2 = st.columns(2)

        with download_col1:
            table_to_csv_download(full_export_df, "agrow_master_registry.csv", key="farmer_tab_master_csv")

        with download_col2:
            table_to_excel_download(full_export_df, "agrow_master_registry.xlsx", key="farmer_tab_master_excel")       

    # =========================================================
    # TAB 4: FARMER VERIFICATION
    # =========================================================
    with tab_verification:
        st.markdown("### Verify an Enumerated Farmer")
        st.caption("Use this workspace during input distribution, field monitoring or farmer support. A valid QR scan opens the same verified record without requiring administrator access.")
        show_public_farmer_verification(compact=True, key_prefix="workspace")

    # =========================================================
    # TAB 5: MARKETLINK
    # =========================================================
    with tab_marketlink:
        render_marketlink_workspace(df, role, user_id)

    # =========================================================
    # TAB 6: ANALYTICS
    # =========================================================
    with tab_analytics:
        c1, c2 = st.columns(2)

        with c1:
            st.subheader("Crop Distribution")
            if not df.empty:
                crop_counts = df["primary_crop"].value_counts().reset_index()
                crop_counts.columns = ["Primary_Crop", "Count"]
                bar_fig = px.bar(crop_counts, x="Primary_Crop", y="Count")
                st.plotly_chart(bar_fig, width="stretch")
            else:
                st.info("No crop data available.")

        with c2:
            st.subheader("NIN Verification Status")
            if not df.empty:
                nin_counts = df["nin_status"].value_counts().reset_index()
                nin_counts.columns = ["NIN_Status", "Count"]
                status_fig = px.bar(nin_counts, x="NIN_Status", y="Count")
                st.plotly_chart(status_fig, width="stretch")
            else:
                st.info("No verification data available.")

        st.subheader("Geographic Farmer Clusters")
        if not df.empty and {"state", "lga", "community", "latitude", "longitude"}.issubset(df.columns):
            cluster_df = df.copy()
            cluster_df["community"] = cluster_df["community"].fillna("Not specified").replace("", "Not specified")
            cluster_summary = (
                cluster_df.groupby(["state", "lga", "community"], dropna=False)
                .agg(
                    Farmers=("farmer_id", "count"),
                    Total_Land_Ha=("farm_size", "sum"),
                    Centre_Latitude=("latitude", "mean"),
                    Centre_Longitude=("longitude", "mean"),
                )
                .reset_index()
                .sort_values(["Farmers", "Total_Land_Ha"], ascending=False)
            )
            st.dataframe(cluster_summary, width="stretch")
            valid_cluster_points = cluster_df[(cluster_df["latitude"] != 0) | (cluster_df["longitude"] != 0)]
            if not valid_cluster_points.empty:
                cluster_map = valid_cluster_points.rename(columns={"latitude": "lat", "longitude": "lon"})
                st.map(cluster_map[["lat", "lon"]])
            st.caption("Use these community clusters to plan input allocation, extension visits, monitoring routes and other farmer support.")
        else:
            st.info("Community cluster analysis will appear after geotagged farmer records are available.")

        if role == "admin":
            st.subheader("👥 Registered Agents")
            agent_df = pd.DataFrame(fetch_all_agents())

            if not agent_df.empty:
                agent_df["State"] = agent_df["State"].apply(normalize_state_name)
                st.dataframe(agent_df, width="stretch")

                a1, a2 = st.columns(2)
                with a1:
                    table_to_csv_download(agent_df, "agrow_registered_agents.csv", key="admin_agents_csv")
                with a2:
                    table_to_excel_download(agent_df, "agrow_registered_agents.xlsx", key="admin_agents_excel")
            else:
                st.info("No registered agents yet.")

    st.markdown(
        '<div class="footer">© DataDev Limited | AGROW — Agricultural Geographic Registration & Operations Workspace v4.4 Farmer Market Portal</div>',
        unsafe_allow_html=True,
    )